from rest_framework import viewsets, status, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied
from roles.permissions import HasModelPermission
from rules import ROLE_ADMIN
from .models import Department, EmployeeDocument, LetterTemplate
from common.bitrix_client import BitrixClient, BitrixEmployeeMock
from .serializers import DepartmentSerializer, EmployeeSerializer, EmployeeDocumentSerializer, LetterTemplateSerializer
from .services import generate_offer_letter, generate_appointment_letter, generate_bond_letter
import os
import mimetypes
from django.http import FileResponse, Http404
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import BaseAuthentication
from rest_framework.exceptions import AuthenticationFailed
from rest_framework_simplejwt.authentication import JWTAuthentication
def sync_employee_to_dict(synced_emp):
    """
    Converts a local SyncedEmployee model instance to a normalized dictionary
    matching the structure returned by BitrixClient._normalize_user.
    """
    dob_str = str(synced_emp.dob) if synced_emp.dob else ''
    joining_date_str = str(synced_emp.joining_date) if synced_emp.joining_date else ''
    
    dept_id = 1
    if synced_emp.department_name:
        try:
            from .models import Department
            dept = Department.objects.filter(name=synced_emp.department_name).first()
            if dept:
                dept_id = dept.id
        except Exception:
            pass

    extra = synced_emp.extra_data or {}

    return {
        'id': str(synced_emp.bitrix_user_id),
        'emp_id': f"LOCAL-{synced_emp.pk}",
        'first_name': synced_emp.first_name,
        'last_name': synced_emp.last_name or '',
        'name': f"{synced_emp.first_name} {synced_emp.last_name or ''}".strip(),
        'email': synced_emp.email or '',
        'work_email': synced_emp.email or '',
        'personal_email': extra.get('personal_email', ''),
        'phone': synced_emp.phone or '',
        'alternate_phone': extra.get('alternate_phone', ''),
        'designation': synced_emp.designation or '',
        'department': dept_id,
        'department_name': synced_emp.department_name or 'Engineering',
        'dob': dob_str,
        'joining_date': joining_date_str,
        'gender': synced_emp.gender or 'Male',
        'address_line1': extra.get('address_line1', ''),
        'address_line2': extra.get('address_line2', ''),
        'city': extra.get('city', ''),
        'state': extra.get('state', ''),
        'pin_code': extra.get('pin_code', ''),
        'employment_type': extra.get('employment_type', 'Full Time'),
        'emergency_contact_name': extra.get('emergency_contact_name', ''),
        'emergency_relationship': extra.get('emergency_relationship', ''),
        'emergency_phone': extra.get('emergency_phone', ''),
        'profile_photo': extra.get('profile_photo', ''),
        'status': synced_emp.status or 'Active',
        'bond_period_months': int(extra.get('bond_period_months', 0)),
        'notice_period_days': int(extra.get('notice_period_days', 30)),
        'bitrix_contact_id': '',
        'bank_account': extra.get('bank_account', ''),
        'bank_name': extra.get('bank_name', ''),
        'aadhaar_masked': extra.get('aadhaar_masked', ''),
        'pan_masked': extra.get('pan_masked', ''),
        'pan_no': extra.get('pan_no', ''),
        'onboarding_complete': synced_emp.onboarding_complete,
        'bitrix_sync_status': getattr(synced_emp, 'bitrix_sync_status', 'Pending'),
        'bitrix_sync_error': getattr(synced_emp, 'bitrix_sync_error', ''),
    }

class QueryParamJWTAuthentication(JWTAuthentication):
    def authenticate(self, request):
        # Check header first using parent implementation
        header_auth = super().authenticate(request)
        if header_auth is not None:
            return header_auth
            
        # Check query parameter
        raw_token = request.query_params.get('token')
        if not raw_token:
            return None
            
        try:
            validated_token = self.get_validated_token(raw_token)
            return self.get_user(validated_token), validated_token
        except Exception:
            raise AuthenticationFailed("Invalid or expired token.")

class SecureDocumentServeView(APIView):
    authentication_classes = [QueryParamJWTAuthentication]
    permission_classes = [IsAuthenticated]


    def get(self, request, emp_id, filename):
        user = request.user
        is_authorized = user.is_superuser or (user.role and user.role.code in ['ADMIN', 'HR'])
        if not is_authorized:
            raise PermissionDenied("You do not have permission to access employee documents.")

        # Resolve path
        file_path = os.path.join(settings.MEDIA_ROOT, 'employees', emp_id, 'docs', filename)
        normalized_path = os.path.normpath(file_path)
        
        # Prevent directory traversal
        if not normalized_path.startswith(os.path.normpath(settings.MEDIA_ROOT)):
            raise PermissionDenied("Access denied.")

        if not os.path.exists(normalized_path) or os.path.isdir(normalized_path):
            raise Http404("Document not found.")

        mime_type, _ = mimetypes.guess_type(normalized_path)
        if not mime_type:
            mime_type = 'application/octet-stream'

        return FileResponse(open(normalized_path, 'rb'), content_type=mime_type)

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all().order_by('name')
    serializer_class = DepartmentSerializer
    permission_classes = [HasModelPermission]

from rest_framework.pagination import PageNumberPagination

class EmployeePagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class EmployeeViewSet(viewsets.ModelViewSet):
    serializer_class = EmployeeSerializer
    permission_classes = [HasModelPermission]
    pagination_class = EmployeePagination

    def get_queryset(self):
        users = BitrixClient.get_all_users()
        mocks = [BitrixEmployeeMock(u) for u in users]
        
        # Load local-only employees from SyncedEmployee
        try:
            from .models import SyncedEmployee
            local_employees = SyncedEmployee.objects.filter(bitrix_user_id__startswith='LOCAL-')
            for local_emp in local_employees:
                local_dict = sync_employee_to_dict(local_emp)
                mocks.append(BitrixEmployeeMock(local_dict))
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"Error loading local employees in get_queryset: {e}")
        
        # Decorate mock users with local exit_request_id if an exit request exists
        try:
            from exit_formality.models import ExitRequest
            exit_requests = ExitRequest.objects.all().order_by('initiated_at')
            exit_requests_map = {str(er.bitrix_user_id): er for er in exit_requests}
            for u in mocks:
                bitrix_id_str = str(u.bitrix_id)
                if bitrix_id_str in exit_requests_map:
                    er = exit_requests_map[bitrix_id_str]
                    u.exit_request_id = er.id
                    u._data['exit_request_id'] = er.id
                    u.exit_request_status = er.status
                    u._data['exit_request_status'] = er.status
        except Exception:
            pass
            
        return mocks

    def get_object(self):
        pk = self.kwargs.get('pk')
        
        # Check if local employee exists first
        try:
            from .models import SyncedEmployee
            lookup_id = pk
            synced_emp = SyncedEmployee.objects.filter(bitrix_user_id=lookup_id).first()
            if not synced_emp and not str(lookup_id).startswith('LOCAL-'):
                synced_emp = SyncedEmployee.objects.filter(bitrix_user_id=f"LOCAL-{lookup_id}").first()
            if synced_emp:
                mock = BitrixEmployeeMock(sync_employee_to_dict(synced_emp))
                self.decorate_mock_with_exit_request(mock)
                return mock
        except Exception:
            pass

        user = BitrixClient.get_user_detail(pk)
        if not user:
            raise Http404("Employee not found in Bitrix24.")
        mock = BitrixEmployeeMock(user)
        self.decorate_mock_with_exit_request(mock)
        return mock

    def decorate_mock_with_exit_request(self, mock):
        try:
            from exit_formality.models import ExitRequest
            exit_req = ExitRequest.objects.filter(bitrix_user_id=str(mock.bitrix_id)).order_by('-initiated_at').first()
            if exit_req:
                mock.exit_request_id = exit_req.id
                mock._data['exit_request_id'] = exit_req.id
                mock.exit_request_status = exit_req.status
                mock._data['exit_request_status'] = exit_req.status
        except Exception:
            pass

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        
        # 1. Search query
        search_query = request.query_params.get('search')
        if search_query:
            search_query = search_query.lower()
            queryset = [
                u for u in queryset 
                if search_query in u.name.lower() or search_query in u.email.lower() or search_query in u.emp_id.lower()
            ]

        # 2. Department filter
        dept_id = request.query_params.get('department')
        if dept_id:
            queryset = [
                u for u in queryset
                if str(u.department) == str(dept_id) or str(dept_id) in [str(d) for d in getattr(u, 'UF_DEPARTMENT', [])]
            ]

        # 3. Employment Type filter
        emp_type = request.query_params.get('employment_type')
        if emp_type:
            queryset = [
                u for u in queryset
                if u.employment_type.upper().replace(' ', '_').replace('-', '_') == emp_type.upper()
            ]

        # 4. Date range filters
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        if from_date:
            queryset = [u for u in queryset if u.joining_date and str(u.joining_date) >= str(from_date)]
        if to_date:
            queryset = [u for u in queryset if u.joining_date and str(u.joining_date) <= str(to_date)]
            
        # 5. List type split logic
        list_type = request.query_params.get('type')
        import datetime
        today = datetime.date.today()
        
        filtered_users = []
        for u in queryset:
            # Set a new joiner flag (e.g. joined within the last 15 days)
            if u.joining_date and (today - datetime.timedelta(days=15)) <= u.joining_date <= today:
                u._data['is_new_joiner'] = True
            else:
                u._data['is_new_joiner'] = False
                
            if list_type == 'onboarding':
                # Employee stays in onboarding on their exact joining date and any day before it
                if not u.joining_date or u.joining_date >= today:
                    # Filter out exited users from onboarding list
                    if u.status != 'Exited':
                        filtered_users.append(u)
            elif list_type == 'all':
                # Active directory only shows people whose joining date is in the past (joining_date < today)
                if u.status != 'Exited':
                    if u.joining_date and u.joining_date < today:
                        filtered_users.append(u)
            elif list_type == 'offboarding':
                is_exited = u.status == 'Exited'
                has_pending_exit = getattr(u, 'exit_request_status', None) == 'PENDING'
                
                if has_pending_exit and not is_exited:
                    filtered_users.append(u)
            elif list_type == 'dismissed':
                if u.status == 'Exited' or getattr(u, 'is_deleted', False) is True:
                    filtered_users.append(u)
            else:
                filtered_users.append(u)

        # 6. Sort ordering
        sort_by = request.query_params.get('sort')
        if sort_by == 'name':
            filtered_users.sort(key=lambda u: u.name.lower())
        elif sort_by == 'date':
            filtered_users.sort(key=lambda u: str(u.joining_date) if u.joining_date else '')
        elif sort_by == 'id':
            filtered_users.sort(key=lambda u: u.emp_id.lower())
                 
        # Pagination
        no_pagination = request.query_params.get('no_pagination') == 'true'
        if not no_pagination:
            page = self.paginate_queryset(filtered_users)
            if page is not None:
                serializer = self.get_serializer([u._data for u in page], many=True)
                return self.get_paginated_response(serializer.data)
            
        serializer = self.get_serializer([u._data for u in filtered_users], many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        employee = self.get_object()
        serializer = self.get_serializer(employee._data)
        return Response(serializer.data)
    def create(self, request, *args, **kwargs):
        data = request.data
        
        # 1. Create a local SyncedEmployee
        import uuid
        from .models import SyncedEmployee
        
        dob_val = data.get('dob') or None
        joining_date_val = data.get('joining_date') or None
        if not dob_val:
            dob_val = None
        if not joining_date_val:
            joining_date_val = None
            
        dept_name = 'Engineering'
        if data.get('department'):
            try:
                from .models import Department
                dept = Department.objects.filter(id=int(data.get('department'))).first()
                if dept:
                    dept_name = dept.name
            except Exception:
                pass

        # Masking Aadhaar & PAN
        aadhaar_val = data.get('aadhaar') or ''
        pan_val = data.get('pan') or ''
        
        def mask_value(val, show_len=4):
            if not val:
                return ""
            val_str = str(val).strip()
            if len(val_str) > show_len:
                return "X" * (len(val_str) - show_len) + val_str[-show_len:]
            return val_str

        extra_data_dict = {
            'personal_email': data.get('personal_email') or '',
            'alternate_phone': data.get('alternate_phone') or '',
            'address_line1': data.get('address_line1') or '',
            'address_line2': data.get('address_line2') or '',
            'city': data.get('city') or '',
            'state': data.get('state') or '',
            'pin_code': data.get('pin_code') or '',
            'employment_type': data.get('employment_type') or 'Full Time',
            'emergency_contact_name': data.get('emergency_contact_name') or '',
            'emergency_relationship': data.get('emergency_relationship') or '',
            'emergency_phone': data.get('emergency_phone') or '',
            'bond_period_months': int(data.get('bond_period_months') or 0),
            'notice_period_days': int(data.get('notice_period_days') or 30),
            'aadhaar_masked': mask_value(aadhaar_val, 4),
            'pan_masked': mask_value(pan_val, 4),
            'pan_no': mask_value(pan_val, 4),
        }

        try:
            synced_emp = SyncedEmployee.objects.create(
                bitrix_user_id=f"LOCAL-{uuid.uuid4().hex[:12]}",
                first_name=data.get('first_name'),
                last_name=data.get('last_name') or '',
                email=data.get('work_email') or data.get('email') or '',
                phone=data.get('phone') or '',
                designation=data.get('designation') or '',
                department_name=dept_name,
                gender=data.get('gender') or '',
                dob=dob_val,
                joining_date=joining_date_val,
                status='Active',
                onboarding_complete=False,
                bitrix_sync_status='Pending',
                extra_data=extra_data_dict
            )
            synced_emp.bitrix_user_id = f"LOCAL-{synced_emp.pk}"
            synced_emp.save()

            from audit_logs.signals import log_action
            if request.user.is_authenticated:
                log_action(request.user, "CREATE", None, f"Created local onboarding employee profile {synced_emp.first_name} {synced_emp.last_name} (Local ID: {synced_emp.pk}).")

            mock_user = sync_employee_to_dict(synced_emp)

            # Trigger welcome email task automatically on creation ONLY IF both joining date and email are present
            if synced_emp.email and synced_emp.joining_date:
                try:
                    from .tasks import send_onboarding_welcome_email
                    send_onboarding_welcome_email.delay(mock_user)
                except Exception:
                    pass

            return Response(mock_user, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    def update(self, request, pk=None, *args, **kwargs):
        data = request.data
        
        # Check if local-only
        is_local = False
        synced_emp = None
        try:
            from .models import SyncedEmployee
            lookup_id = pk
            synced_emp = SyncedEmployee.objects.filter(bitrix_user_id=lookup_id).first()
            if not synced_emp and not str(lookup_id).startswith('LOCAL-'):
                synced_emp = SyncedEmployee.objects.filter(bitrix_user_id=f"LOCAL-{lookup_id}").first()
            if synced_emp:
                is_local = True
        except Exception:
            pass

        if is_local:
            dob_val = data.get('dob') or None
            joining_date_val = data.get('joining_date') or None
            if not dob_val:
                dob_val = None
            if not joining_date_val:
                joining_date_val = None

            dept_name = synced_emp.department_name
            if data.get('department'):
                try:
                    from .models import Department
                    dept = Department.objects.filter(id=int(data.get('department'))).first()
                    if dept:
                        dept_name = dept.name
                except Exception:
                    pass

            try:
                synced_emp.first_name = data.get('first_name', synced_emp.first_name)
                synced_emp.last_name = data.get('last_name', synced_emp.last_name)
                synced_emp.email = data.get('work_email') or data.get('email') or synced_emp.email
                synced_emp.phone = data.get('phone', synced_emp.phone)
                synced_emp.designation = data.get('designation', synced_emp.designation)
                synced_emp.department_name = dept_name
                synced_emp.gender = data.get('gender', synced_emp.gender)
                if dob_val:
                    synced_emp.dob = dob_val
                if joining_date_val:
                    synced_emp.joining_date = joining_date_val
                
                # Update extra_data
                extra = synced_emp.extra_data or {}
                if 'personal_email' in data:
                    extra['personal_email'] = data.get('personal_email')
                if 'alternate_phone' in data:
                    extra['alternate_phone'] = data.get('alternate_phone')
                if 'address_line1' in data:
                    extra['address_line1'] = data.get('address_line1')
                if 'address_line2' in data:
                    extra['address_line2'] = data.get('address_line2')
                if 'city' in data:
                    extra['city'] = data.get('city')
                if 'state' in data:
                    extra['state'] = data.get('state')
                if 'pin_code' in data:
                    extra['pin_code'] = data.get('pin_code')
                if 'employment_type' in data:
                    extra['employment_type'] = data.get('employment_type')
                if 'emergency_contact_name' in data:
                    extra['emergency_contact_name'] = data.get('emergency_contact_name')
                if 'emergency_relationship' in data:
                    extra['emergency_relationship'] = data.get('emergency_relationship')
                if 'emergency_phone' in data:
                    extra['emergency_phone'] = data.get('emergency_phone')
                if 'bond_period_months' in data:
                    extra['bond_period_months'] = int(data.get('bond_period_months') or 0)
                if 'notice_period_days' in data:
                    extra['notice_period_days'] = int(data.get('notice_period_days') or 30)

                def mask_value(val, show_len=4):
                    if not val:
                        return ""
                    val_str = str(val).strip()
                    if len(val_str) > show_len:
                        return "X" * (len(val_str) - show_len) + val_str[-show_len:]
                    return val_str
                
                if 'aadhaar' in data:
                    extra['aadhaar_masked'] = mask_value(data.get('aadhaar'), 4)
                if 'pan' in data:
                    extra['pan_masked'] = mask_value(data.get('pan'), 4)
                    extra['pan_no'] = mask_value(data.get('pan'), 4)
                
                # Bank details
                bank_account = data.get('bank_account')
                bank_name = data.get('bank_name')
                if bank_account is not None:
                    extra['bank_account'] = bank_account
                if bank_name is not None:
                    extra['bank_name'] = bank_name

                synced_emp.extra_data = extra
                synced_emp.save()

                if bank_account is not None or bank_name is not None:
                    from salary.models import EmployeeBankDetail
                    detail, _ = EmployeeBankDetail.objects.get_or_create(bitrix_user_id=synced_emp.bitrix_user_id)
                    if bank_account is not None:
                        detail.bank_account_no = bank_account
                    if bank_name is not None:
                        detail.bank_name = bank_name
                    detail.save()

                from audit_logs.signals import log_action
                if request.user.is_authenticated:
                    log_action(request.user, "UPDATE", None, f"Updated local employee details (Local ID: {pk}).")

                mock_user = sync_employee_to_dict(synced_emp)
                return Response(mock_user)
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        # Fallback to Bitrix update
        webhook = BitrixClient.get_webhook_url()
        gender_code = 'M'
        if data.get('gender') == 'FEMALE':
            gender_code = 'F'
        elif data.get('gender') == 'OTHER':
            gender_code = 'O'
            
        payload = {
            'ID': pk,
            'NAME': data.get('first_name'),
            'LAST_NAME': data.get('last_name'),
            'EMAIL': data.get('work_email') or data.get('email'),
            'PERSONAL_MOBILE': data.get('phone'),
            'WORK_POSITION': data.get('designation'),
            'PERSONAL_BIRTHDAY': data.get('dob'),
            'PERSONAL_CITY': data.get('city') or data.get('address_line1'),
            'PERSONAL_STATE': data.get('state'),
            'PERSONAL_ZIP': data.get('pin_code'),
            'PERSONAL_GENDER': gender_code,
            'UF_PERSONAL_EMAIL': data.get('personal_email'),
            'PERSONAL_MAILBOX': data.get('personal_email')
        }
        
        try:
            import requests
            res = requests.post(f"{webhook}/user.update.json", json=payload, timeout=10)
            
            if not res.ok:
                update_url_crm = f"{webhook}/crm.contact.update"
                emails_crm = []
                email_val = data.get('work_email') or data.get('email')
                if email_val:
                    emails_crm.append({'VALUE': email_val, 'VALUE_TYPE': 'WORK'})
                if data.get('personal_email'):
                    emails_crm.append({'VALUE': data.get('personal_email'), 'VALUE_TYPE': 'HOME'})
                    
                payload_crm = {
                    'id': pk,
                    'fields': {
                        'NAME': data.get('first_name'),
                        'LAST_NAME': data.get('last_name'),
                        'EMAIL': emails_crm,
                        'PHONE': [{'VALUE': data.get('phone'), 'VALUE_TYPE': 'WORK'}],
                        'POST': data.get('designation'),
                        'UF_PERSONAL_EMAIL': data.get('personal_email'),
                        'PERSONAL_MAILBOX': data.get('personal_email')
                    }
                }
                res = requests.post(update_url_crm, json=payload_crm, timeout=10)
                
            if res.ok:
                bank_account = data.get('bank_account')
                bank_name = data.get('bank_name')
                if bank_account is not None or bank_name is not None:
                    from salary.models import EmployeeBankDetail
                    detail, _ = EmployeeBankDetail.objects.get_or_create(bitrix_user_id=pk)
                    if bank_account is not None:
                        detail.bank_account_no = bank_account
                    if bank_name is not None:
                        detail.bank_name = bank_name
                    detail.save()

                BitrixClient.get_all_users(force_refresh=True)
                user = BitrixClient.get_user_detail(pk)
                
                try:
                    from notifications.models import BitrixSyncLog
                    BitrixSyncLog.objects.create(
                        employee_id=pk,
                        employee_name=user.get('name') if user else f"{data.get('first_name')} {data.get('last_name')}".strip(),
                        action_type='Contact Update',
                        status='SUCCESS',
                        retry_count=0
                    )
                except Exception:
                    pass

                from audit_logs.signals import log_action
                if request.user.is_authenticated:
                    log_action(request.user, "UPDATE", None, f"Updated Bitrix24 employee details (ID: {pk}).")
                return Response(user)
            else:
                emp_name = f"{data.get('first_name')} {data.get('last_name')}".strip()
                try:
                    from notifications.models import BitrixSyncLog, Notification
                    from django.contrib.auth import get_user_model
                    User = get_user_model()
                    
                    BitrixSyncLog.objects.create(
                        employee_id=pk,
                        employee_name=emp_name,
                        action_type='Contact Update',
                        status='FAILED',
                        retry_count=0,
                        error_message=res.text
                    )
                    
                    admins = User.objects.filter(role__code='ADMIN')
                    notif_msg = f"Bitrix24 sync failed for {emp_name} (Contact Update). Retry available."
                    for admin in admins:
                        Notification.objects.create(
                            recipient=admin,
                            notif_type='ERROR',
                            message=notif_msg,
                            link="/admin/bitrix/sync-log/"
                        )
                except Exception:
                    pass

                return Response({'error': f"Bitrix24 API Error: {res.text}"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None, *args, **kwargs):
        return self.update(request, pk, *args, **kwargs)

    def destroy(self, request, pk=None, *args, **kwargs):
        # Check if local-only
        is_local = False
        synced_emp = None
        try:
            from .models import SyncedEmployee
            lookup_id = pk
            synced_emp = SyncedEmployee.objects.filter(bitrix_user_id=lookup_id).first()
            if not synced_emp and not str(lookup_id).startswith('LOCAL-'):
                synced_emp = SyncedEmployee.objects.filter(bitrix_user_id=f"LOCAL-{lookup_id}").first()
            if synced_emp:
                is_local = True
        except Exception:
            pass

        if is_local:
            try:
                local_id = synced_emp.bitrix_user_id
                synced_emp.delete()
                # Delete related local documents/salary structures/bank details
                from employee_onboarding.models import EmployeeDocument
                EmployeeDocument.objects.filter(bitrix_user_id=local_id).delete()
                
                from salary.models import SalaryStructure, EmployeeBankDetail
                SalaryStructure.objects.filter(bitrix_user_id=local_id).delete()
                EmployeeBankDetail.objects.filter(bitrix_user_id=local_id).delete()
                
                return Response(status=status.HTTP_204_NO_CONTENT)
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        # Fallback to Bitrix delete
        webhook = BitrixClient.get_webhook_url()
        delete_url = f"{webhook}/crm.contact.delete"
        try:
            import requests
            requests.post(delete_url, json={'id': pk}, timeout=10)
        except Exception:
            pass
        BitrixClient.get_all_users(force_refresh=True)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['POST'], url_path='invite-to-bitrix')
    def invite_to_bitrix(self, request, pk=None):
        from .models import SyncedEmployee
        lookup_id = pk
        synced_emp = SyncedEmployee.objects.filter(bitrix_user_id=lookup_id).first()
        if not synced_emp and not str(lookup_id).startswith('LOCAL-'):
            synced_emp = SyncedEmployee.objects.filter(bitrix_user_id=f"LOCAL-{lookup_id}").first()
            
        if not synced_emp:
            return Response({'error': 'Local employee not found.'}, status=status.HTTP_404_NOT_FOUND)
        
        if synced_emp.bitrix_sync_status == 'Synced':
            return Response({'error': 'Employee is already synced to Bitrix.'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if email is present
        if not synced_emp.email:
            return Response({'error': 'Cannot invite to Bitrix: Email is missing.'}, status=status.HTTP_400_BAD_REQUEST)
            
        webhook = BitrixClient.get_webhook_url()
        
        # Map gender
        gender_code = 'M'
        if synced_emp.gender == 'Female':
            gender_code = 'F'
        elif synced_emp.gender == 'Other':
            gender_code = 'O'
            
        # Map department
        dept_list = [1]
        if synced_emp.department_name:
            try:
                from .models import Department
                dept = Department.objects.filter(name=synced_emp.department_name).first()
                if dept:
                    dept_list = [dept.id]
            except Exception:
                pass
                
        user_payload = {
            'EMAIL': synced_emp.email,
            'NAME': synced_emp.first_name,
            'LAST_NAME': synced_emp.last_name or '',
            'PERSONAL_MOBILE': synced_emp.phone or '',
            'WORK_POSITION': synced_emp.designation or '',
            'UF_DEPARTMENT': dept_list,
            'PERSONAL_BIRTHDAY': str(synced_emp.dob) if synced_emp.dob else '',
            'PERSONAL_GENDER': gender_code,
        }
        
        try:
            import requests
            res = requests.post(f"{webhook}/user.add.json", json=user_payload, timeout=10)
            
            if not res.ok:
                emails = [{'VALUE': synced_emp.email, 'VALUE_TYPE': 'WORK'}]
                crm_payload = {
                    'fields': {
                        'NAME': synced_emp.first_name,
                        'LAST_NAME': synced_emp.last_name or '',
                        'EMAIL': emails,
                        'PHONE': [{'VALUE': synced_emp.phone or '', 'VALUE_TYPE': 'WORK'}],
                        'POST': synced_emp.designation or '',
                        'UF_ONBOARDING_STATUS': 'Pending',
                        'UF_CRM_ONBOARDING_STATUS': 'Pending',
                    }
                }
                res = requests.post(f"{webhook}/crm.contact.add", json=crm_payload, timeout=10)
                
            if res.ok:
                contact_id = str(res.json().get('result'))
                
                # Update references in all tables
                old_local_id = synced_emp.bitrix_user_id
                
                from employee_onboarding.models import EmployeeDocument
                EmployeeDocument.objects.filter(bitrix_user_id=old_local_id).update(bitrix_user_id=contact_id)
                
                from salary.models import SalaryStructure
                SalaryStructure.objects.filter(bitrix_user_id=old_local_id).update(bitrix_user_id=contact_id)
                
                from salary.models import EmployeeBankDetail
                EmployeeBankDetail.objects.filter(bitrix_user_id=old_local_id).update(bitrix_user_id=contact_id)
                
                synced_emp.bitrix_user_id = contact_id
                synced_emp.bitrix_sync_status = 'Synced'
                synced_emp.bitrix_sync_error = None
                synced_emp.save()
                
                # Force refresh cache
                BitrixClient.get_all_users(force_refresh=True)
                
                # Log success in BitrixSyncLog
                try:
                    from notifications.models import BitrixSyncLog
                    BitrixSyncLog.objects.create(
                        employee_id=contact_id,
                        employee_name=f"{synced_emp.first_name} {synced_emp.last_name}".strip(),
                        action_type='Contact Sync',
                        status='SUCCESS',
                        retry_count=0
                    )
                except Exception:
                    pass
                    
                # Trigger welcome email task automatically now that they are synced and details are complete!
                try:
                    mock_user = BitrixClient.get_user_detail(contact_id)
                    from .tasks import send_onboarding_welcome_email
                    send_onboarding_welcome_email.delay(mock_user)
                except Exception:
                    pass
                    
                return Response({
                    'success': True,
                    'message': 'Employee invited to Bitrix successfully.',
                    'new_id': contact_id
                })
            else:
                error_msg = res.text
                synced_emp.bitrix_sync_status = 'Failed'
                synced_emp.bitrix_sync_error = error_msg
                synced_emp.save()
                
                return Response({
                    'success': False,
                    'error': f"Bitrix24 API Error: {error_msg}"
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['POST'], url_path='send-welcome-email')
    def send_welcome_email(self, request, pk=None):
        employee = self.get_object()
        
        email_val = employee.get('email') or employee.get('work_email') or employee.get('personal_email')
        joining_date_val = employee.get('joining_date')
        
        if not email_val or not joining_date_val:
            return Response({
                'error': 'Cannot send welcome email. Joining date and email must both be provided.'
            }, status=status.HTTP_400_BAD_REQUEST)
            
        from .tasks import send_onboarding_welcome_email
        emp_data = employee._data if hasattr(employee, '_data') else dict(employee)
        send_onboarding_welcome_email.delay(emp_data)
        
        return Response({'message': 'Welcome email sent successfully.'})

    @action(detail=True, methods=['POST'], url_path='manual-graduate')
    def manual_graduate(self, request, pk=None):
        from .models import SyncedEmployee
        lookup_id = pk
        synced_emp = SyncedEmployee.objects.filter(bitrix_user_id=lookup_id).first()
        if not synced_emp and not str(lookup_id).startswith('LOCAL-'):
            synced_emp = SyncedEmployee.objects.filter(bitrix_user_id=f"LOCAL-{lookup_id}").first()
            
        if synced_emp:
            synced_emp.onboarding_complete = True
            synced_emp.save()
        else:
            webhook = BitrixClient.get_webhook_url()
            update_url = f"{webhook}/crm.contact.update"
            payload = {
                'id': pk,
                'fields': {
                    'UF_ONBOARDING_STATUS': 'Completed',
                    'UF_CRM_ONBOARDING_STATUS': 'Completed'
                }
            }
            try:
                import requests
                requests.post(update_url, json=payload, timeout=10)
            except Exception:
                pass
        BitrixClient.get_all_users(force_refresh=True)
        return Response({'message': 'Employee onboarding completed successfully.'})

    @action(detail=False, methods=['POST'], url_path='bitrix-webhook', authentication_classes=[], permission_classes=[])
    def bitrix_webhook(self, request):
        from .models import SyncedEmployee
        data = request.data
        
        # Determine the payload structure
        emp_data = data
        
        # If it's a standard Bitrix webhook payload with event
        if 'event' in data and data['event'] in ['ONCRMCONTACTADD', 'ONCRMCONTACTUPDATE', 'ONUSERADD', 'ONUSERUPDATE']:
            contact_id = data.get('data', {}).get('FIELDS', {}).get('ID')
            if contact_id:
                user_detail = BitrixClient.get_user_detail(contact_id)
                if user_detail:
                    emp_data = user_detail
                else:
                    return Response({'status': 'Contact not found'}, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({'status': 'No ID provided'}, status=status.HTTP_400_BAD_REQUEST)

        # Extract data from the payload or the fetched user details
        bitrix_id = str(emp_data.get('ID') or emp_data.get('id') or emp_data.get('bitrix_id') or data.get('id') or '')
        
        if not bitrix_id:
            import uuid
            bitrix_id = f"LOCAL-{uuid.uuid4().hex[:12]}"
            
        full_name = emp_data.get('NAME') or emp_data.get('name') or data.get('name') or ''
        first_name = emp_data.get('first_name') or data.get('first_name') or ''
        last_name = emp_data.get('LAST_NAME') or emp_data.get('last_name') or data.get('last_name') or ''
        
        # If full name is provided but no first/last name, split it
        if full_name and not first_name:
            parts = full_name.split(' ', 1)
            first_name = parts[0]
            if len(parts) > 1:
                last_name = parts[1]
                
        if not first_name:
            first_name = "Unknown"

        email = emp_data.get('EMAIL') or emp_data.get('email') or emp_data.get('work_email') or data.get('email') or ''
        phone = emp_data.get('PERSONAL_MOBILE') or emp_data.get('phone') or data.get('phone', '')
        designation = emp_data.get('WORK_POSITION') or emp_data.get('designation') or data.get('designation', '')
        department_name = emp_data.get('department_name') or data.get('department_name', '')
        
        synced_emp, created = SyncedEmployee.objects.get_or_create(
            bitrix_user_id=bitrix_id,
            defaults={
                'first_name': first_name,
                'last_name': last_name,
                'email': email,
                'phone': phone,
                'designation': designation,
                'department_name': department_name,
                'bitrix_sync_status': 'Synced',
            }
        )
        
        if not created:
            synced_emp.first_name = first_name or synced_emp.first_name
            synced_emp.last_name = last_name or synced_emp.last_name
            synced_emp.email = email or synced_emp.email
            synced_emp.phone = phone or synced_emp.phone
            synced_emp.designation = designation or synced_emp.designation
            if department_name:
                synced_emp.department_name = department_name
            synced_emp.bitrix_sync_status = 'Synced'
            synced_emp.save()
            
        return Response({'status': 'success', 'message': 'Employee data received and stored in local DB. Sync to Bitrix skipped.'})

    @action(detail=True, methods=['POST'], url_path='retry-bitrix-sync')
    def retry_bitrix_sync(self, request, pk=None):
        return Response({'message': 'Project is running in Pure API mode. Sync is always active.'})

    @action(detail=True, methods=['POST'], url_path='generate-offer-letter')
    def generate_offer_letter_api(self, request, pk=None):
        employee = self.get_object()
        custom_context = request.data.get('custom_context', None)
        try:
            doc = generate_offer_letter(employee, user=request.user, custom_context=custom_context)
            # Sync to Bitrix24 timeline
            from .tasks import attach_document_to_bitrix24_timeline, send_offer_letter_email
            attach_document_to_bitrix24_timeline.delay(doc.id)
            send_offer_letter_email.delay(employee.id, doc.id)
            return Response({
                'message': 'Offer letter generated successfully.',
                'document': EmployeeDocumentSerializer(doc).data
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['POST'], url_path='generate-appointment-letter')
    def generate_appointment_letter_api(self, request, pk=None):
        employee = self.get_object()
        custom_context = request.data.get('custom_context', None)
        try:
            doc = generate_appointment_letter(employee, user=request.user, custom_context=custom_context)
            # Sync to Bitrix24 timeline
            from .tasks import attach_document_to_bitrix24_timeline
            attach_document_to_bitrix24_timeline.delay(doc.id)
            return Response({
                'message': 'Appointment letter generated successfully.',
                'document': EmployeeDocumentSerializer(doc).data
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['POST'], url_path='generate-bond-letter')
    def generate_bond_letter_api(self, request, pk=None):
        employee = self.get_object()
        custom_context = request.data.get('custom_context', None)
        
        bond_period = employee.bond_period_months
        if custom_context and 'bond_period_months' in custom_context:
            try:
                bond_period = int(custom_context['bond_period_months'] or 0)
            except ValueError:
                pass
                
        if bond_period <= 0:
            return Response({
                'error': 'This employee does not have a bond period specified.'
            }, status=status.HTTP_400_BAD_REQUEST)
        try:
            doc = generate_bond_letter(employee, user=request.user, custom_context=custom_context)
            # Sync to Bitrix24 timeline
            from .tasks import attach_document_to_bitrix24_timeline
            attach_document_to_bitrix24_timeline.delay(doc.id)
            return Response({
                'message': 'Bond letter generated successfully.',
                'document': EmployeeDocumentSerializer(doc).data
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['POST'], url_path='preview-letter')
    def preview_letter_api(self, request, pk=None):
        employee = self.get_object()
        doc_type = request.data.get('doc_type', None)
        custom_context = request.data.get('custom_context', None)
        
        if not doc_type:
            return Response({'error': 'doc_type is required'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            from .services import render_letter_to_html
            html_string = render_letter_to_html(employee, doc_type, custom_context)
            return Response({'html': html_string})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['POST'], url_path='bitrix-webhook', permission_classes=[], authentication_classes=[])
    def bitrix_webhook(self, request):
        data = request.data or {}
        bitrix_id = None
        
        # 1. Check direct keys for ID (including crm.id)
        for key in ['id', 'ID', 'bitrix_id', 'bitrix_user_id', 'crm.id']:
            if key in data:
                bitrix_id = data[key]
                break
                
        # 2. Check nested inside 'data' key
        if not bitrix_id and isinstance(data.get('data'), dict):
            nested_data = data['data']
            if isinstance(nested_data.get('FIELDS'), dict):
                fields = nested_data['FIELDS']
                for key in ['ID', 'id', 'crm.id']:
                    if key in fields:
                        bitrix_id = fields[key]
                        break
            if not bitrix_id:
                for key in ['id', 'ID', 'bitrix_id', 'bitrix_user_id', 'crm.id']:
                    if key in nested_data:
                        bitrix_id = nested_data[key]
                        break

        # 3. Check nested inside 'fields' key
        if not bitrix_id and isinstance(data.get('fields'), dict):
            fields = data['fields']
            for key in ['ID', 'id', 'crm.id']:
                if key in fields:
                    bitrix_id = fields[key]
                    break

        first_name = data.get('first_name') or data.get('name')
        email = data.get('email') or data.get('work_email')

        if bitrix_id:
            from .tasks import process_bitrix_webhook_task
            process_bitrix_webhook_task.delay(data)
            return Response({
                'status': 'queued',
                'message': f'Sync queued for employee ID {bitrix_id}.'
            }, status=status.HTTP_202_ACCEPTED)

        elif first_name and email:
            from .tasks import process_bitrix_webhook_task
            process_bitrix_webhook_task.delay(data)
            return Response({
                'status': 'queued',
                'message': f'Creation and sync queued for {first_name}.'
            }, status=status.HTTP_202_ACCEPTED)

        return Response({'error': 'Missing ID or required fields (first_name, email).'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['POST'], url_path='import-excel')
    def excel_import(self, request):
        user = request.user
        is_admin = user.is_superuser or (user.role and user.role.code == 'ADMIN')
        if not is_admin:
            raise PermissionDenied("Only Admin can import employee details.")

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'file is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Failed to parse Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        headers = [str(ws.cell(row=1, column=col).value).strip() for col in range(1, ws.max_column + 1)]
        header_map = {}
        for idx, h in enumerate(headers):
            header_map[h.lower()] = idx + 1

        def get_val(row, header_name, default=""):
            idx = header_map.get(header_name.lower())
            if idx is None:
                return default
            val = ws.cell(row=row, column=idx).value
            return val if val is not None else default

        import datetime
        def parse_date(val):
            if not val:
                return None
            if isinstance(val, (datetime.date, datetime.datetime)):
                return val.date() if isinstance(val, datetime.datetime) else val
            try:
                for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y'):
                    try:
                        return datetime.datetime.strptime(str(val).strip(), fmt).date()
                    except ValueError:
                        continue
            except Exception:
                pass
            return None

        from decimal import Decimal
        def parse_decimal(val):
            if val is None or str(val).strip() == "":
                return Decimal("0.00")
            try:
                cleaned = str(val).strip().replace("$", "").replace(",", "")
                return Decimal(cleaned)
            except Exception:
                return Decimal("0.00")

        total_records = 0
        success_count = 0
        failed_count = 0
        failed_rows_data = []

        from django.db import transaction
        from salary.models import EmployeeBankDetail, SalaryStructure

        # Fetch all active users from Bitrix
        bitrix_users = BitrixClient.get_all_users()

        with transaction.atomic():
            for row_idx in range(2, ws.max_row + 1):
                first_name = str(get_val(row_idx, 'First Name')).strip()
                last_name = str(get_val(row_idx, 'Last Name')).strip()
                if not first_name and not last_name:
                    continue

                total_records += 1
                row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
                sid = transaction.savepoint()

                try:
                    parsed_emp_id = str(get_val(row_idx, 'Employee ID')).strip()
                    parsed_email = str(get_val(row_idx, 'Email')).strip()
                    parsed_phone = str(get_val(row_idx, 'Phone')).strip()

                    if not parsed_email:
                        raise ValueError("Email is required")

                    # Find matching employee in bitrix_users
                    employee = None
                    for u in bitrix_users:
                        mock_emp = BitrixEmployeeMock(u)
                        if parsed_emp_id and mock_emp.emp_id.lower() == parsed_emp_id.lower():
                            employee = mock_emp
                            break
                        if parsed_email and mock_emp.email.lower() == parsed_email.lower():
                            employee = mock_emp
                            break
                        if parsed_phone and mock_emp.phone == parsed_phone:
                            employee = mock_emp
                            break
                        if mock_emp.name.lower() == f"{first_name} {last_name}".lower().strip():
                            employee = mock_emp
                            break

                    gender_val = str(get_val(row_idx, 'Gender')).strip().upper()
                    gender_code = 'M'
                    if gender_val in ('FEMALE', 'F'):
                        gender_code = 'F'
                    elif gender_val in ('OTHER', 'O'):
                        gender_code = 'O'

                    dept_val = get_val(row_idx, 'Department')
                    dept_list = [1]
                    if dept_val:
                        try:
                            dept_list = [int(float(str(dept_val)))]
                        except ValueError:
                            dept_obj = Department.objects.filter(name__iexact=str(dept_val).strip()).first()
                            if dept_obj:
                                dept_list = [dept_obj.id]

                    webhook = BitrixClient.get_webhook_url()
                    contact_id = None

                    if employee:
                        contact_id = employee.bitrix_id
                        # Update employee details in Bitrix24
                        payload = {
                            'ID': contact_id,
                            'NAME': first_name,
                            'LAST_NAME': last_name,
                            'EMAIL': parsed_email,
                            'PERSONAL_MOBILE': parsed_phone,
                            'WORK_POSITION': str(get_val(row_idx, 'Designation')).strip(),
                            'PERSONAL_BIRTHDAY': str(get_val(row_idx, 'DOB')).strip(),
                            'PERSONAL_CITY': str(get_val(row_idx, 'Address Line 1')).strip(),
                            'PERSONAL_STATE': str(get_val(row_idx, 'State')).strip(),
                            'PERSONAL_ZIP': str(get_val(row_idx, 'PIN Code')).strip(),
                            'PERSONAL_GENDER': gender_code,
                        }
                        
                        import requests
                        res = requests.post(f"{webhook}/user.update.json", json=payload, timeout=10)
                        if not res.ok:
                            payload_crm = {
                                'id': contact_id,
                                'fields': {
                                    'NAME': first_name,
                                    'LAST_NAME': last_name,
                                    'EMAIL': [{'VALUE': parsed_email, 'VALUE_TYPE': 'WORK'}],
                                    'PHONE': [{'VALUE': parsed_phone, 'VALUE_TYPE': 'WORK'}],
                                    'POST': str(get_val(row_idx, 'Designation')).strip(),
                                }
                            }
                            requests.post(f"{webhook}/crm.contact.update", json=payload_crm, timeout=10)
                    else:
                        # Create new contact in Bitrix24
                        user_payload = {
                            'EMAIL': parsed_email,
                            'NAME': first_name,
                            'LAST_NAME': last_name,
                            'PERSONAL_MOBILE': parsed_phone,
                            'WORK_POSITION': str(get_val(row_idx, 'Designation')).strip(),
                            'UF_DEPARTMENT': dept_list,
                            'PERSONAL_BIRTHDAY': str(get_val(row_idx, 'DOB')).strip(),
                            'PERSONAL_GENDER': gender_code,
                        }
                        
                        import requests
                        res = requests.post(f"{webhook}/user.add.json", json=user_payload, timeout=10)
                        if res.ok:
                            contact_id = str(res.json().get('result'))
                        else:
                            crm_payload = {
                                'fields': {
                                    'NAME': first_name,
                                    'LAST_NAME': last_name,
                                    'EMAIL': [{'VALUE': parsed_email, 'VALUE_TYPE': 'WORK'}],
                                    'PHONE': [{'VALUE': parsed_phone, 'VALUE_TYPE': 'WORK'}],
                                    'POST': str(get_val(row_idx, 'Designation')).strip(),
                                    'UF_ONBOARDING_STATUS': 'Pending',
                                    'UF_CRM_ONBOARDING_STATUS': 'Pending',
                                }
                            }
                            res_crm = requests.post(f"{webhook}/crm.contact.add", json=crm_payload, timeout=10)
                            if res_crm.ok:
                                contact_id = str(res_crm.json().get('result'))

                        if not contact_id:
                            # Fallback mock ID
                            import uuid
                            contact_id = f"MOCK-{uuid.uuid4().hex[:8]}"

                    # Save or update Bank details
                    bank_acc = str(get_val(row_idx, 'Bank Account Number')).strip()
                    bank_name = str(get_val(row_idx, 'Bank Name')).strip()
                    if bank_acc or bank_name:
                        EmployeeBankDetail.objects.update_or_create(
                            bitrix_user_id=contact_id,
                            defaults={
                                'bank_account_no': bank_acc,
                                'bank_name': bank_name
                            }
                        )

                    # Save or update Salary Structure
                    gross_salary = parse_decimal(get_val(row_idx, 'Gross Salary'))
                    pf_contribution = parse_decimal(get_val(row_idx, 'PF Contribution'))
                    esi = parse_decimal(get_val(row_idx, 'ESI'))
                    labour_welfare_fund = parse_decimal(get_val(row_idx, 'Labour Welfare Fund'))
                    professional_tax = parse_decimal(get_val(row_idx, 'Professional Tax'))
                    other_deductions = parse_decimal(get_val(row_idx, 'Other Deductions'))
                    
                    salary_eff_from = parse_date(get_val(row_idx, 'Salary Effective From'))
                    joining_date = parse_date(get_val(row_idx, 'Joining Date'))
                    
                    effective_from = salary_eff_from or joining_date or datetime.date.today()

                    SalaryStructure.objects.update_or_create(
                        bitrix_user_id=contact_id,
                        effective_from=effective_from,
                        defaults={
                            'gross_salary': gross_salary,
                            'pf_contribution': pf_contribution,
                            'esi': esi,
                            'labour_welfare_fund': labour_welfare_fund,
                            'professional_tax': professional_tax,
                            'other_deductions': other_deductions
                        }
                    )

                    transaction.savepoint_commit(sid)
                    success_count += 1
                except Exception as e:
                    transaction.savepoint_rollback(sid)
                    failed_count += 1
                    failed_rows_data.append((row_vals, str(e)))

        if failed_count > 0:
            import openpyxl
            from openpyxl import Workbook
            from django.http import HttpResponse
            
            error_wb = Workbook()
            error_ws = error_wb.active
            error_ws.title = "Import Errors"
            error_ws.append(headers + ["Errors"])
            
            for row_vals, err_msg in failed_rows_data:
                padded_vals = list(row_vals) + [""] * (len(headers) - len(row_vals))
                error_ws.append(padded_vals + [err_msg])
                
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="employee_import_errors.xlsx"'
            error_wb.save(response)
            return response

        # Force refresh Bitrix users cache
        BitrixClient.get_all_users(force_refresh=True)

        return Response({
            'message': f'Successfully imported {success_count} employee records.'
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['GET'], url_path='export-excel')
    def excel_export(self, request):
        queryset = self.get_queryset()
        
        dept_id = request.query_params.get('department')
        emp_type = request.query_params.get('employment_type')
        status_param = request.query_params.get('status')
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        
        if emp_type:
            queryset = [emp for emp in queryset if emp.employment_type == emp_type]
        if status_param and status_param != 'All':
            queryset = [emp for emp in queryset if emp.status == status_param]
        if from_date:
            try:
                import datetime
                from_dt = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
                queryset = [emp for emp in queryset if emp.joining_date >= from_dt]
            except ValueError:
                pass
        if to_date:
            try:
                import datetime
                to_dt = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
                queryset = [emp for emp in queryset if emp.joining_date <= to_dt]
            except ValueError:
                pass

        import openpyxl
        from openpyxl import Workbook
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        headers = [
            'Employee ID', 'First Name', 'Last Name', 'Email', 'Phone', 'Alternate Phone',
            'DOB', 'Gender', 'Address Line 1', 'Address Line 2', 'City', 'State', 'PIN Code',
            'Department', 'Designation', 'Employment Type', 'Joining Date', 'Notice Period', 'Bond Period',
            'Emergency Name', 'Emergency Relationship', 'Emergency Phone', 'Status', 'Onboarding Status',
            'Bank Account Number', 'Bank Name', 'Gross Salary', 'PF Contribution', 'ESI',
            'Labour Welfare Fund', 'Professional Tax', 'Other Deductions', 'Salary Effective From'
        ]
        ws.append(headers)

        from salary.models import EmployeeBankDetail, SalaryStructure

        for emp in queryset:
            detail = EmployeeBankDetail.objects.filter(bitrix_user_id=emp.bitrix_id).first()
            bank_acc = detail.bank_account_no if detail else ""
            bank_nm = detail.bank_name if detail else ""
            
            struct = SalaryStructure.objects.filter(bitrix_user_id=emp.bitrix_id).order_by('-effective_from').first()
            gross_val = float(struct.gross_salary) if struct else 0.0
            pf_val = float(struct.pf_contribution) if struct else 0.0
            esi_val = float(struct.esi) if struct else 0.0
            lwf_val = float(struct.labour_welfare_fund) if struct else 0.0
            pt_val = float(struct.professional_tax) if struct else 0.0
            other_val = float(struct.other_deductions) if struct else 0.0
            eff_val = struct.effective_from.strftime('%Y-%m-%d') if (struct and struct.effective_from) else ""

            ws.append([
                emp.emp_id,
                emp.first_name,
                emp.last_name,
                emp.email,
                emp.phone,
                emp.alternate_phone or '',
                emp.dob.strftime('%Y-%m-%d') if emp.dob else '',
                emp.get_gender_display(),
                emp.address_line1,
                emp.address_line2 or '',
                emp.city,
                emp.get_state_display(),
                emp.pin_code,
                emp.department_name,
                emp.designation,
                emp.get_employment_type_display(),
                emp.joining_date.strftime('%Y-%m-%d') if emp.joining_date else '',
                emp.notice_period_days,
                emp.bond_period_months,
                emp.emergency_contact_name,
                emp.get_emergency_relationship_display(),
                emp.emergency_phone,
                emp.status,
                'Completed' if emp.onboarding_complete else 'Under Onboarding',
                bank_acc,
                bank_nm,
                gross_val,
                pf_val,
                esi_val,
                lwf_val,
                pt_val,
                other_val,
                eff_val
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="employees_export.xlsx"'
        wb.save(response)
        return response


class EmployeeDocumentViewSet(viewsets.ModelViewSet):
    serializer_class = EmployeeDocumentSerializer
    permission_classes = [HasModelPermission]

    def get_queryset(self):
        queryset = EmployeeDocument.objects.all().order_by('-upload_date')
        employee_id = self.request.query_params.get('employee_id') or self.request.query_params.get('bitrix_user_id')
        if employee_id:
            queryset = queryset.filter(bitrix_user_id=employee_id)
        return queryset

    def perform_create(self, serializer):
        doc = serializer.save()
        # Log upload in audit logs
        from audit_logs.signals import log_action
        log_action(self.request.user, "UPLOAD", doc, f"Uploaded document '{doc.get_doc_type_display()}' ({doc.original_filename}) for employee {doc.bitrix_user_id}.")
        # Sync to Bitrix24 timeline
        from .tasks import attach_document_to_bitrix24_timeline
        attach_document_to_bitrix24_timeline.delay(doc.id)


from .services import DEFAULT_TEMPLATES


class LetterTemplateViewSet(viewsets.ModelViewSet):
    serializer_class = LetterTemplateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Automatically seed templates if they do not exist
        for t in DEFAULT_TEMPLATES:
            if not LetterTemplate.objects.filter(name=t['name']).exists():
                LetterTemplate.objects.create(**t)
                    
        # Automatically seed exit templates if they do not exist
        exit_templates_seeding = [
            ('RELIEVING_LETTER', 'Relieving Letter Template', 'pdf_relieving_letter.html'),
            ('EXPERIENCE_LETTER', 'Experience Letter Template', 'pdf_experience_letter.html'),
            ('NOTICE_LETTER', 'Notice Period Letter Template', 'pdf_notice_letter.html'),
            ('NOC_LETTER', 'NOC Letter Template', 'pdf_noc_letter.html'),
            ('FF_SETTLEMENT_LETTER', 'F&F Settlement Letter Template', 'pdf_ff_settlement_letter.html'),
            ('FF_SALARY_SLIP', 'Final Month Payslip Template', 'pdf_ff_salary_slip.html'),
        ]
        
        for name, title, filename in exit_templates_seeding:
            if not LetterTemplate.objects.filter(name=name).exists():
                try:
                    file_path = os.path.join(settings.BASE_DIR, 'templates', 'exit', filename)
                    with open(file_path, 'r', encoding='utf-8') as f:
                        html_content = f.read()
                    LetterTemplate.objects.create(
                        name=name,
                        title=title,
                        html_content=html_content,
                        allow_hr_edit=False
                    )
                except Exception as e:
                    pass
                    
        return LetterTemplate.objects.all().order_by('name')


    def update(self, request, *args, **kwargs):
        obj = self.get_object()
        user = request.user
        is_admin = user.is_superuser or (user.role and user.role.code == 'ADMIN')
        
        if not is_admin:
            user_perms = list(user.role.permissions.values_list('codename', flat=True)) if user.role else []
            is_hr = user.role and user.role.code == 'HR'
            
            is_exit_template = obj.name in [
                'RELIEVING_LETTER', 'EXPERIENCE_LETTER', 'NOTICE_LETTER', 
                'NOC_LETTER', 'FF_SETTLEMENT_LETTER', 'FF_SALARY_SLIP'
            ]
            
            if is_exit_template:
                has_manage_perm = 'exit.manage_templates' in user_perms
                if not has_manage_perm:
                    raise PermissionDenied("You do not have permission to edit this exit template.")
            else:
                has_manage_perm = 'onboarding.manage_templates' in user_perms
                if not has_manage_perm:
                    if not is_hr or not obj.allow_hr_edit:
                        raise PermissionDenied("You do not have permission to edit this template.")
                
        return super().update(request, *args, **kwargs)


from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render

@csrf_exempt
def employees_hybrid_view(request, *args, **kwargs):
    """
    Custom hybrid view for the /employees/ endpoint.
    Handles GET: lists employees (renders layout or returns API response).
    Handles POST:
      - If it's a frontend AJAX request (Accept: application/json or X-Requested-With: XMLHttpRequest),
        routes to 'create' to add a user to Bitrix.
      - Otherwise (e.g. from Bitrix webhook), routes to 'bitrix_webhook' to sync the user.
    """
    from .views import EmployeeViewSet
    
    is_json = (
        'application/json' in request.META.get('HTTP_ACCEPT', '') or
        request.headers.get('x-requested-with') == 'XMLHttpRequest' or
        request.GET.get('format') == 'json'
    )
    
    if request.method == 'GET':
        if is_json:
            api_view_func = EmployeeViewSet.as_view({'get': 'list'})
            response = api_view_func(request, *args, **kwargs)
            if hasattr(response, 'render'):
                response.render()
            return response
        else:
            path = request.path.rstrip('/')
            initial_tab = 'onboarding'
            if path.endswith('active'):
                initial_tab = 'all'
            elif path.endswith('offboarding'):
                initial_tab = 'offboarding'
            elif path.endswith('dismissed'):
                initial_tab = 'dismissed'
                
            return render(request, 'base/layout.html', {
                'initial_onboarding_tab': initial_tab
            })
    elif request.method == 'POST':
        is_ajax_post = (
            'application/json' in request.META.get('HTTP_ACCEPT', '') or
            request.headers.get('x-requested-with') == 'XMLHttpRequest'
        )
        
        if is_ajax_post:
            import json
            is_webhook = False
            try:
                body_data = json.loads(request.body)
                if isinstance(body_data, dict) and ('event' in body_data or 'data' in body_data):
                    is_webhook = True
            except Exception:
                pass
            
            if is_webhook:
                api_view_func = EmployeeViewSet.as_view({'post': 'bitrix_webhook'}, permission_classes=[], authentication_classes=[])
            else:
                api_view_func = EmployeeViewSet.as_view({'post': 'create'})
        else:
            api_view_func = EmployeeViewSet.as_view({'post': 'bitrix_webhook'}, permission_classes=[], authentication_classes=[])
            
        response = api_view_func(request, *args, **kwargs)
        if hasattr(response, 'render'):
            response.render()
        return response

class BitrixDataReceiveAPIView(APIView):
    """
    Standalone API specifically to receive Bitrix employee data 
    and save it locally without syncing back to Bitrix.
    """
    authentication_classes = []
    permission_classes = []

    def post(self, request, *args, **kwargs):
        from .models import SyncedEmployee
        data = request.data
        
        emp_data = data
        
        # Check if it's an event payload
        if 'event' in data and data['event'] in ['ONCRMCONTACTADD', 'ONCRMCONTACTUPDATE', 'ONUSERADD', 'ONUSERUPDATE']:
            contact_id = data.get('data', {}).get('FIELDS', {}).get('ID')
            if contact_id:
                user_detail = BitrixClient.get_user_detail(contact_id)
                if user_detail:
                    emp_data = user_detail
                else:
                    return Response({'status': 'Contact not found'}, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({'status': 'No ID provided'}, status=status.HTTP_400_BAD_REQUEST)

        bitrix_id = str(emp_data.get('ID') or emp_data.get('id') or emp_data.get('bitrix_id') or data.get('id') or '')
        # Always prefix with LOCAL- to ensure it shows up in employee list
        if not bitrix_id or not bitrix_id.startswith('LOCAL-'):
            import uuid
            if bitrix_id:
                bitrix_id = f"LOCAL-{bitrix_id}"
            else:
                bitrix_id = f"LOCAL-{uuid.uuid4().hex[:12]}"
            
        full_name = emp_data.get('NAME') or emp_data.get('name') or data.get('name') or ''
        first_name = emp_data.get('first_name') or data.get('first_name') or ''
        last_name = emp_data.get('LAST_NAME') or emp_data.get('last_name') or data.get('last_name') or ''
        
        # If full name is provided but no first/last name, split it
        if full_name and not first_name:
            parts = full_name.split(' ', 1)
            first_name = parts[0]
            if len(parts) > 1:
                last_name = parts[1]
                
        if not first_name:
            first_name = "Unknown"

        email = emp_data.get('EMAIL') or emp_data.get('email') or emp_data.get('work_email') or data.get('email') or ''
        phone = emp_data.get('PERSONAL_MOBILE') or emp_data.get('phone') or data.get('phone', '')
        designation = emp_data.get('WORK_POSITION') or emp_data.get('designation') or data.get('designation', '')
        department_name = emp_data.get('department_name') or data.get('department_name', '')
        
        synced_emp, created = SyncedEmployee.objects.get_or_create(
            bitrix_user_id=bitrix_id,
            defaults={
                'first_name': first_name,
                'last_name': last_name,
                'email': email,
                'phone': phone,
                'designation': designation,
                'department_name': department_name,
                'bitrix_sync_status': 'Synced',
            }
        )
        
        if not created:
            synced_emp.first_name = first_name or synced_emp.first_name
            synced_emp.last_name = last_name or synced_emp.last_name
            synced_emp.email = email or synced_emp.email
            synced_emp.phone = phone or synced_emp.phone
            synced_emp.designation = designation or synced_emp.designation
            if department_name:
                synced_emp.department_name = department_name
            synced_emp.bitrix_sync_status = 'Synced'
            synced_emp.save()
            
        return Response({'status': 'success', 'message': 'Standalone API: Employee data received and stored in local DB. Sync to Bitrix skipped.'})
