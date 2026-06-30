# apps/salary/tests.py
import datetime
import io
import openpyxl
from decimal import Decimal
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework import status
from rest_framework.test import APITestCase
from unittest.mock import patch

# Python 3.14 compatibility patch for django template context copy
from django.template.context import BaseContext
def patch_copy(self):
    dup = self.__class__()
    dup.dicts = self.dicts[:]
    return dup
BaseContext.__copy__ = patch_copy

from roles.models import Role
from employee_onboarding.models import Department
from salary.models import SalaryStructure, SalarySlip, SalaryImportBatch
from common.bitrix_client import BitrixEmployeeMock

User = get_user_model()

class SalaryModuleTests(APITestCase):

    def setUp(self):
        from django.core.cache import cache
        cache.clear()
        # Setup mock patchers
        self.get_all_users_patcher = patch('common.bitrix_client.BitrixClient.get_all_users')
        self.mock_get_all = self.get_all_users_patcher.start()
        
        self.get_user_detail_patcher = patch('common.bitrix_client.BitrixClient.get_user_detail')
        self.mock_get_detail = self.get_user_detail_patcher.start()

        user_dict = {
            'id': '10',
            'emp_id': 'BITRIX-10',
            'first_name': 'John',
            'last_name': 'Doe',
            'name': 'John Doe',
            'email': 'emp@test.com',
            'phone': '9876543210',
            'designation': 'Software Engineer',
            'department_name': 'Engineering',
            'dob': '1995-05-10',
            'gender': 'Male',
            'joining_date': '2026-06-01',
            'status': 'Active'
        }
        self.mock_get_all.return_value = [user_dict]
        self.mock_get_detail.return_value = user_dict

        self.employee = BitrixEmployeeMock(user_dict)

        # Create roles
        self.admin_role = Role.objects.create(name="Admin", code="ADMIN")
        self.hr_role = Role.objects.create(name="HR", code="HR")
        self.employee_role = Role.objects.create(name="Employee", code="EMPLOYEE")

        # Create users
        self.admin_user = User.objects.create_superuser(
            username="admin@test.com", email="admin@test.com", password="password", role=self.admin_role
        )
        self.hr_user = User.objects.create_user(
            username="hr@test.com", email="hr@test.com", password="password", role=self.hr_role
        )
        self.employee_user = User.objects.create_user(
            username="emp@test.com", email="emp@test.com", password="password", role=self.employee_role
        )

        # Create department
        self.dept = Department.objects.create(name="Engineering")

        # Create salary structure
        self.structure = SalaryStructure.objects.create(
            bitrix_user_id=self.employee.bitrix_id,
            gross_salary=Decimal('87000.00'),
            pf_contribution=Decimal('6000.00'),
            esi=Decimal('0.00'),
            labour_welfare_fund=Decimal('0.00'),
            professional_tax=Decimal('200.00'),
            other_deductions=Decimal('0.00'),
            effective_from=datetime.date(2026, 6, 1)
        )

    def tearDown(self):
        self.get_all_users_patcher.stop()
        self.get_user_detail_patcher.stop()

    def test_admin_salary_export_template(self):
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('salary_export') + "?month=6&year=2026"
        res = self.client.get(url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res['Content-Type'], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Load exported sheet
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=1).value, 1)
        self.assertEqual(ws.cell(row=2, column=2).value, self.employee.name)

    def test_unauthorized_salary_export_fails(self):
        self.client.force_authenticate(user=self.hr_user)
        url = reverse('salary_export') + "?month=6&year=2026"
        res = self.client.get(url)
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

        # Employee cannot export template
        self.client.force_authenticate(user=self.employee_user)
        res = self.client.get(url)
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

    def create_excel_file(self, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            "Sr. No.", "Name", "Designation", "Month days", "Worked days", 
            "Weekend", "CL", "Extra", "Payable Days", "Month Salary", 
            "Payable Salary", "Extra days working", "Fine/Advance", "Net Payable", 
            "Bank A/c No.", "Bank"
        ]
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def test_admin_salary_import_success(self):
        self.client.force_authenticate(user=self.admin_user)
        excel_content = self.create_excel_file([
            [1, "John Doe", "Software Engineer", 30.0, 26.0, 4.0, 0.0, 0.0, 30.0, 87000.0, 87000.0, 0.0, 0.0, 87000.0, "123456", "ICICI"]
        ])
        
        uploaded_file = SimpleUploadedFile("salary.xlsx", excel_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        url = reverse('salary_import')
        data = {
            'file': uploaded_file,
            'month': 6,
            'year': 2026
        }
        res = self.client.post(url, data, format='multipart')
        if res.status_code != 200:
            print("IMPORT SUCCESS TEST FAIL DATA:", res.data)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['success'], 1)
        self.assertEqual(res.data['failed'], 0)

        # Check slip is created in DB as draft
        slip = SalarySlip.objects.get(bitrix_user_id=self.employee.bitrix_id, month=6, year=2026)
        self.assertEqual(slip.status, 'draft')
        self.assertEqual(slip.month_salary, Decimal('87000.00'))
        self.assertEqual(slip.fine_advance, Decimal('0.00'))
        self.assertEqual(slip.net_payable, Decimal('87000.00'))
        self.assertTrue(slip.pdf_file)

    def test_admin_salary_import_partial_failure(self):
        self.client.force_authenticate(user=self.admin_user)
        excel_content = self.create_excel_file([
            # Success row
            [1, "John Doe", "Software Engineer", 30.0, 26.0, 4.0, 0.0, 0.0, 30.0, 87000.0, 87000.0, 0.0, 0.0, 87000.0, "123456", "ICICI"],
            # Failure row: invalid Name
            [2, "Invalid Name", "Design", 30.0, 26.0, 4.0, 0.0, 0.0, 30.0, 40000.0, 40000.0, 0.0, 0.0, 40000.0, "654321", "HDFC"]
        ])
        
        uploaded_file = SimpleUploadedFile("salary_partial.xlsx", excel_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        url = reverse('salary_import')
        data = {
            'file': uploaded_file,
            'month': 6,
            'year': 2026
        }
        res = self.client.post(url, data, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['success'], 1)
        self.assertEqual(res.data['failed'], 1)
        self.assertIsNotNone(res.data['error_report_url'])

        # Verify batch status is partial
        batch = SalaryImportBatch.objects.first()
        self.assertEqual(batch.status, 'partial')

    def test_admin_salary_import_total_failure_rollback(self):
        self.client.force_authenticate(user=self.admin_user)
        excel_content = self.create_excel_file([
            # Failure row 1: invalid Name
            [1, "Invalid Name", "Design", 30.0, 26.0, 4.0, 0.0, 0.0, 30.0, 40000.0, 40000.0, 0.0, 0.0, 40000.0, "654321", "HDFC"],
            # Failure row 2: invalid gross/month salary amount (-500)
            [2, "John Doe", "Software Engineer", 30.0, 26.0, 4.0, 0.0, 0.0, 30.0, -500.0, -500.0, 0.0, 0.0, -500.0, "123456", "ICICI"]
        ])

        uploaded_file = SimpleUploadedFile("salary_failed.xlsx", excel_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        url = reverse('salary_import')
        data = {
            'file': uploaded_file,
            'month': 6,
            'year': 2026
        }
        res = self.client.post(url, data, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(res.data['failed'], 2)

        # Verify no slips created at all
        self.assertEqual(SalarySlip.objects.count(), 0)

    def test_publish_workflow(self):
        # Create draft slip
        slip = SalarySlip.objects.create(
            bitrix_user_id=self.employee.bitrix_id,
            month=6,
            year=2026,
            month_salary=87000.00,
            status='draft'
        )
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('salary_publish')
        res = self.client.post(url, {'month': 6, 'year': 2026}, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        
        # Verify slip is published
        slip.refresh_from_db()
        self.assertEqual(slip.status, 'published')

    def test_salary_history_scoping(self):
        # Create slips
        slip1 = SalarySlip.objects.create(
            bitrix_user_id=self.employee.bitrix_id, month=5, year=2026, month_salary=87000.00, status='published'
        )
        slip2 = SalarySlip.objects.create(
            bitrix_user_id=self.employee.bitrix_id, month=6, year=2026, month_salary=90000.00, status='draft'
        )

        # Test employee request
        self.client.force_authenticate(user=self.employee_user)
        url = reverse('salary_history')
        res = self.client.get(url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        # Employees should only see their own published records (slip1)
        self.assertEqual(res.data['count'], 1)
        self.assertEqual(res.data['results'][0]['month'], 5)

        # Test HR request (can see draft too)
        self.client.force_authenticate(user=self.hr_user)
        res = self.client.get(url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['count'], 2)

    def test_salary_slip_pdf_and_zip_download(self):
        # Create multiple published slips
        slip1 = SalarySlip.objects.create(
            bitrix_user_id=self.employee.bitrix_id, month=4, year=2026, month_salary=87000.00, status='published'
        )
        slip2 = SalarySlip.objects.create(
            bitrix_user_id=self.employee.bitrix_id, month=5, year=2026, month_salary=87000.00, status='published'
        )

        # Download single month
        self.client.force_authenticate(user=self.employee_user)
        url = reverse('salary_download') + "?type=single&month=4&year=2026"
        res = self.client.get(url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res['Content-Type'], "application/pdf")

        # Download multi-month (last3) should zip
        url = reverse('salary_download') + "?type=last3"
        res = self.client.get(url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res['Content-Type'], "application/zip")

    def test_bank_details_autosave_and_prefill(self):
        # Update mock values to fully-populated user dictionaries to prevent serializer KeyError
        user_dict_complete = {
            'id': '10',
            'emp_id': 'BITRIX-10',
            'first_name': 'John',
            'last_name': 'Doe',
            'name': 'John Doe',
            'email': 'emp@test.com',
            'work_email': 'emp@test.com',
            'personal_email': '',
            'phone': '9876543210',
            'designation': 'Software Engineer',
            'department_name': 'Engineering',
            'dob': '1995-05-10',
            'gender': 'Male',
            'joining_date': '2026-06-01',
            'status': 'Active',
            'onboarding_complete': True,
            'alternate_phone': '',
            'address_line1': 'Mohali',
            'address_line2': '',
            'city': 'Mohali',
            'state': 'Punjab',
            'pin_code': '160055',
            'employment_type': 'Full Time',
            'emergency_contact_name': 'Emergency Contact',
            'emergency_relationship': 'Friend',
            'emergency_phone': '9876543210'
        }
        self.mock_get_detail.return_value = user_dict_complete
        self.mock_get_all.return_value = [user_dict_complete]

        from salary.models import EmployeeBankDetail
        # Verify no bank details exist initially
        self.assertEqual(EmployeeBankDetail.objects.count(), 0)

        # 1. Test Auto-save on Excel Import
        self.client.force_authenticate(user=self.admin_user)
        excel_content = self.create_excel_file([
            [1, "John Doe", "Software Engineer", 30.0, 26.0, 4.0, 0.0, 0.0, 30.0, 87000.0, 87000.0, 0.0, 0.0, 87000.0, "987654321", "MockBank"]
        ])
        uploaded_file = SimpleUploadedFile("salary.xlsx", excel_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        url = reverse('salary_import')
        res = self.client.post(url, {'file': uploaded_file, 'month': 6, 'year': 2026}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

        # Verify EmployeeBankDetail is saved
        bank_detail = EmployeeBankDetail.objects.get(bitrix_user_id=self.employee.bitrix_id)
        self.assertEqual(bank_detail.bank_account_no, "987654321")
        self.assertEqual(bank_detail.bank_name, "MockBank")

        # 2. Test Prefill on Excel Import when Excel is blank
        excel_content_blank = self.create_excel_file([
            [1, "John Doe", "Software Engineer", 30.0, 26.0, 4.0, 0.0, 0.0, 30.0, 87000.0, 87000.0, 0.0, 0.0, 87000.0, "", ""]
        ])
        uploaded_file_blank = SimpleUploadedFile("salary_blank.xlsx", excel_content_blank, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        # Let's delete the old slip to test creation
        SalarySlip.objects.all().delete()
        res = self.client.post(url, {'file': uploaded_file_blank, 'month': 6, 'year': 2026}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

        # Verify slip has the prefilled bank account/name from database
        slip = SalarySlip.objects.get(bitrix_user_id=self.employee.bitrix_id, month=6, year=2026)
        self.assertEqual(slip.bank_account_no, "987654321")
        self.assertEqual(slip.bank_name, "MockBank")

        # 3. Test update on Salary slip edit
        edit_url = reverse('salary_edit', kwargs={'pk': slip.id})
        res = self.client.put(edit_url, {'bank_account_no': '111222333', 'bank_name': 'NewMockBank'}, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

        # Verify persistent EmployeeBankDetail is updated
        bank_detail.refresh_from_db()
        self.assertEqual(bank_detail.bank_account_no, "111222333")
        self.assertEqual(bank_detail.bank_name, "NewMockBank")

        # 4. Test Prefill in Employee Details API & view
        emp_detail_url = reverse('employees_detail_view', kwargs={'pk': int(self.employee.bitrix_id)})
        self.client.force_authenticate(user=self.admin_user)
        res = self.client.get(emp_detail_url, {'format': 'json'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['bank_account'], "111222333")
        self.assertEqual(res.data['bank_name'], "NewMockBank")

        # 5. Test update via Employee details PATCH request
        res = self.client.patch(emp_detail_url + "?format=json", {'bank_account': '555666777', 'bank_name': 'FinalMockBank'}, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

        # Verify updated bank details are saved to database
        bank_detail.refresh_from_db()
        self.assertEqual(bank_detail.bank_account_no, "555666777")
        self.assertEqual(bank_detail.bank_name, "FinalMockBank")

    @patch('exit_formality.tasks.send_exit_documents_after_fully_exited')
    @patch('exit_formality.tasks.update_bitrix24_on_exit')
    @patch('exit_formality.tasks.send_exit_initiation_email')
    def test_exited_employee_payroll(self, mock_send_init, mock_update_bitrix, mock_send_docs):
        from exit_formality.models import ExitRequest

        # 1. Mock exited employee in Bitrix
        exited_user_dict = {
            'id': '20',
            'emp_id': 'BITRIX-20',
            'first_name': 'Jane',
            'last_name': 'Smith',
            'name': 'Jane Smith',
            'email': 'jane@test.com',
            'phone': '9876543211',
            'designation': 'QA Engineer',
            'department_name': 'Engineering',
            'dob': '1996-06-12',
            'gender': 'Female',
            'joining_date': '2026-06-01',
            'status': 'Exited'
        }

        # Keep original mocked users list (from setUp) and add the exited user
        original_mock_users = [self.mock_get_all.return_value[0]]
        self.mock_get_all.return_value = original_mock_users + [exited_user_dict]

        # Let's verify that running export without an active ExitRequest does NOT include Jane Smith
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('salary_export') + "?month=6&year=2026"
        res = self.client.get(url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb.active
        # Rows are: headers (row 1), John Doe (row 2)
        # Verify that there are only 2 rows in the sheet (header + John Doe)
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(row=2, column=2).value, "John Doe")

        # Let's verify that importing a salary slip for Jane Smith without ExitRequest fails
        excel_content = self.create_excel_file([
            [1, "John Doe", "Software Engineer", 30.0, 26.0, 4.0, 0.0, 0.0, 30.0, 87000.0, 87000.0, 0.0, 0.0, 87000.0, "123456", "ICICI"],
            [2, "Jane Smith", "QA Engineer", 30.0, 26.0, 4.0, 0.0, 0.0, 30.0, 60000.0, 60000.0, 0.0, 0.0, 60000.0, "654321", "HDFC"]
        ])
        uploaded_file = SimpleUploadedFile("salary_fail_exited.xlsx", excel_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        url_import = reverse('salary_import')
        data = {
            'file': uploaded_file,
            'month': 6,
            'year': 2026
        }
        res_import = self.client.post(url_import, data, format='multipart')
        self.assertEqual(res_import.status_code, status.HTTP_200_OK)
        self.assertEqual(res_import.data['success'], 1)
        self.assertEqual(res_import.data['failed'], 1)  # Jane Smith should fail

        # Delete any created slips to test empty template export
        SalarySlip.objects.all().delete()

        # Now let's create a pending ExitRequest for Jane Smith
        exit_request = ExitRequest.objects.create(
            bitrix_user_id='20',
            resignation_date=datetime.date(2026, 6, 15),
            last_working_day=datetime.date(2026, 6, 30),
            exit_type='RESIGNATION',
            exit_reason='Career growth',
            mode_of_resignation='Email',
            status='PENDING'
        )

        # Let's create a salary structure for Jane Smith as well
        SalaryStructure.objects.create(
            bitrix_user_id='20',
            gross_salary=Decimal('60000.00'),
            pf_contribution=Decimal('4000.00'),
            esi=Decimal('0.00'),
            labour_welfare_fund=Decimal('0.00'),
            professional_tax=Decimal('200.00'),
            other_deductions=Decimal('0.00'),
            effective_from=datetime.date(2026, 6, 1)
        )

        # Now exporting should include Jane Smith
        res_export2 = self.client.get(url)
        self.assertEqual(res_export2.status_code, status.HTTP_200_OK)
        wb2 = openpyxl.load_workbook(io.BytesIO(res_export2.content))
        ws2 = wb2.active
        # Rows should be: header (row 1), John Doe (row 2), Jane Smith (row 3)
        self.assertEqual(ws2.max_row, 3)
        self.assertEqual(ws2.cell(row=2, column=2).value, "John Doe")
        self.assertEqual(ws2.cell(row=3, column=2).value, "Jane Smith")

        # Now importing should succeed for Jane Smith too
        uploaded_file2 = SimpleUploadedFile("salary_success_exited.xlsx", excel_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        res_import2 = self.client.post(url_import, {
            'file': uploaded_file2,
            'month': 6,
            'year': 2026
        }, format='multipart')
        self.assertEqual(res_import2.status_code, status.HTTP_200_OK)
        self.assertEqual(res_import2.data['success'], 2)
        self.assertEqual(res_import2.data['failed'], 0)

        # Check slip is created in DB for Jane Smith
        self.assertTrue(SalarySlip.objects.filter(bitrix_user_id='20', month=6, year=2026).exists())

        # Let's verify that if the ExitRequest is FULLY_EXITED or CANCELLED, it is NOT included and imports fail
        exit_request.status = 'FULLY_EXITED'
        exit_request.save()

        # Delete any created slips to test empty template export
        SalarySlip.objects.all().delete()

        # Export should NOT include Jane Smith anymore
        res_export3 = self.client.get(url)
        self.assertEqual(res_export3.status_code, status.HTTP_200_OK)
        wb3 = openpyxl.load_workbook(io.BytesIO(res_export3.content))
        ws3 = wb3.active
        self.assertEqual(ws3.max_row, 2)
        self.assertEqual(ws3.cell(row=2, column=2).value, "John Doe")

        # Import should fail for Jane Smith
        uploaded_file3 = SimpleUploadedFile("salary_fail_exited_completed.xlsx", excel_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        # Let's delete the Jane Smith slip to check import failure behavior
        SalarySlip.objects.filter(bitrix_user_id='20', month=6, year=2026).delete()
        res_import3 = self.client.post(url_import, {
            'file': uploaded_file3,
            'month': 6,
            'year': 2026
        }, format='multipart')
        self.assertEqual(res_import3.status_code, status.HTTP_200_OK)
        self.assertEqual(res_import3.data['success'], 1)
        self.assertEqual(res_import3.data['failed'], 1)

    def test_salary_employee_summary(self):
        # Create a couple of slips for the employee
        SalarySlip.objects.create(
            bitrix_user_id=self.employee.bitrix_id,
            month=5,
            year=2026,
            month_salary=Decimal('87000.00'),
            worked_days=Decimal('26.00'),
            month_days=Decimal('30.00'),
            net_payable=Decimal('87000.00'),
            fine_advance=Decimal('500.00'),
            payment_status='paid',
            payment_date=datetime.date(2026, 6, 5),
            status='published'
        )
        
        SalarySlip.objects.create(
            bitrix_user_id=self.employee.bitrix_id,
            month=6,
            year=2026,
            month_salary=Decimal('87000.00'),
            worked_days=Decimal('26.00'),
            month_days=Decimal('30.00'),
            net_payable=Decimal('85000.00'),
            fine_advance=Decimal('2000.00'),
            payment_status='paid',
            payment_date=datetime.date(2026, 7, 5),
            status='published'
        )

        # Authenticate as admin to view the summary
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('salary_employee_summary', kwargs={'employee_id': int(self.employee.bitrix_id)})
        res = self.client.get(url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(Decimal(res.data['total_salary_credited']), Decimal('148300.00')) # Calculated by save(): 74900 + 73400
        self.assertEqual(Decimal(res.data['total_deductions']), Decimal('2500.00')) # 500 + 2000
        self.assertEqual(res.data['total_payslips'], 2)
        self.assertEqual(res.data['last_payment_date'], '2026-07-05')

    def test_import_preserves_custom_values(self):
        # Authenticate as admin
        self.client.force_authenticate(user=self.admin_user)
        
        # Create Excel file where Month Salary is 0, but Payable Salary is 2000 and Net Payable is 20000
        excel_content = self.create_excel_file([
            [1, "John Doe", "Software Engineer", 30.0, 30.0, 3.0, 0.0, 2.0, 0.0, 0.0, 2000.0, 0.0, 0.0, 20000.0, "", ""]
        ])
        uploaded_file = SimpleUploadedFile("salary_custom.xlsx", excel_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        url = reverse('salary_import')
        data = {
            'file': uploaded_file,
            'month': 6,
            'year': 2026
        }
        res = self.client.post(url, data, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['success'], 1)
        self.assertEqual(res.data['failed'], 0)

        # Check that the slip has the exact imported custom values (not overwritten to 0)
        slip = SalarySlip.objects.get(bitrix_user_id=self.employee.bitrix_id, month=6, year=2026)
        self.assertEqual(slip.month_salary, Decimal('0.00'))
        self.assertEqual(slip.payable_salary, Decimal('2000.00'))
        self.assertEqual(slip.net_payable, Decimal('20000.00'))
        # And payable_days is calculated correctly based on attendance even though it was 0 in Excel
        self.assertEqual(slip.payable_days, Decimal('35.00'))

    def test_future_salary_slip_visibility(self):
        # Create a slip in the future (e.g. November 2026, while today is June 2026)
        SalarySlip.objects.create(
            bitrix_user_id=self.employee.bitrix_id,
            month=11,
            year=2026,
            month_salary=Decimal('87000.00'),
            worked_days=Decimal('30.00'),
            month_days=Decimal('30.00'),
            net_payable=Decimal('87000.00'),
            status='published'
        )

        # Authenticate as admin
        self.client.force_authenticate(user=self.admin_user)

        # Fetch history for a range covering November 2026
        url = reverse('salary_history') + f"?employee_id={self.employee.bitrix_id}&from=2026-11&to=2030-12"
        res = self.client.get(url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        
        # Verify that the future slip is successfully returned
        slips = res.data['results'] if 'results' in res.data else res.data
        self.assertEqual(len(slips), 1)
        self.assertEqual(slips[0]['month'], 11)
        self.assertEqual(slips[0]['year'], 2026)

    def test_manual_generation_carry_forward(self):
        # Create a mock batch
        batch = SalaryImportBatch.objects.create(
            month=1,
            year=2026,
            file_name="test.xlsx",
            uploaded_by=self.admin_user,
            status='success'
        )
        # Create a slip in January 2026
        prior_slip = SalarySlip.objects.create(
            bitrix_user_id=self.employee.bitrix_id,
            month=1,
            year=2026,
            month_days=Decimal('31.00'),
            worked_days=Decimal('20.00'),
            weekend=Decimal('8.00'),
            cl=Decimal('2.00'),
            extra=Decimal('1.00'),
            month_salary=Decimal('62000.00'),
            extra_days_working=Decimal('2000.00'),
            fine_advance=Decimal('1000.00'),
            status='published',
            uploaded_batch=batch
        )
        
        # Authenticate as admin
        self.client.force_authenticate(user=self.admin_user)
        
        # Manually generate for February 2026
        url = reverse('salary_generate_individual')
        data = {
            'employee_id': self.employee.bitrix_id,
            'month': 2,
            'year': 2026
        }
        res = self.client.post(url, data, format='json')
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        
        # Verify the new slip is created and has correct copied values for Feb 2026 (28 days)
        slip = SalarySlip.objects.get(bitrix_user_id=self.employee.bitrix_id, month=2, year=2026)
        self.assertEqual(slip.month_days, Decimal('28.00'))
        self.assertEqual(slip.worked_days, Decimal('20.00'))
        # payable_days = 20 + 8 + 2 + 1 = 31
        self.assertEqual(slip.payable_days, Decimal('31.00'))
        # payable_salary is copied exactly from January: 62000.00
        self.assertEqual(slip.payable_salary, Decimal('62000.00'))
        # net_payable is copied exactly from January: 63000.00
        self.assertEqual(slip.net_payable, Decimal('63000.00'))
        self.assertTrue(bool(slip.pdf_file))
        
        # Generate again: should reuse and regenerate without failing
        res2 = self.client.post(url, data, format='json')
        self.assertEqual(res2.status_code, status.HTTP_200_OK)



