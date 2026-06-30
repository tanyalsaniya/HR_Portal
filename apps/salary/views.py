import datetime
import os
import openpyxl
from decimal import Decimal
from django.db import models, transaction
from django.core.files.base import ContentFile
from django.conf import settings
from django.http import HttpResponse, Http404
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import viewsets, status, serializers
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.pagination import PageNumberPagination

from common.bitrix_client import BitrixClient, BitrixEmployeeMock
from .models import SalaryStructure, SalarySlip, SalaryImportBatch, SalaryIncrementReminder, SalaryIncrementApproval
from .serializers import (
    SalaryStructureSerializer, SalarySlipSerializer, SalaryImportBatchSerializer,
    SalaryIncrementReminderSerializer, SalaryIncrementApprovalSerializer
)
from .services import generate_payslip_pdf, generate_increment_letter_pdf, generate_payslips_zip, num_to_words, get_latest_prior_slip

# Helper to check roles
def check_role(user):
    if not user or not user.is_authenticated:
        return None
    if user.is_superuser or (user.role and user.role.code == 'ADMIN'):
        return 'admin'
    if user.role and user.role.code == 'HR':
        return 'hr'
    return 'employee'

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100

# Legacy viewsets for backward compatibility with UI
class SalaryStructureViewSet(viewsets.ModelViewSet):
    serializer_class = SalaryStructureSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = SalaryStructure.objects.all().order_by('-effective_from')
        employee_id = self.request.query_params.get('employee_id')
        if employee_id:
            queryset = queryset.filter(bitrix_user_id=employee_id)
        return queryset

    def perform_create(self, serializer):
        # Admin check
        if check_role(self.request.user) != 'admin':
            raise PermissionDenied("Only Admin can manage salary structures.")
        serializer.save()


class SalaryIncrementViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'approve':
            return SalaryIncrementApprovalSerializer
        return SalaryIncrementReminderSerializer

    def get_queryset(self):
        if self.action in ('list_approvals', 'retrieve_approval'):
            return SalaryIncrementApproval.objects.all().order_by('-approved_at')
        return SalaryIncrementReminder.objects.all().order_by('-anniversary_date')

    @action(detail=False, methods=['GET'], url_path='approvals')
    def list_approvals(self, request):
        queryset = SalaryIncrementApproval.objects.all().order_by('-approved_at')
        serializer = SalaryIncrementApprovalSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['POST'], url_path='approve')
    @transaction.atomic
    def approve(self, request, pk=None):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can approve increments.")

        reminder = get_object_or_404(SalaryIncrementReminder, pk=pk)
        if reminder.status == 'Actioned':
            return Response({'error': 'This increment reminder has already been actioned.'}, status=status.HTTP_400_BAD_REQUEST)

        bitrix_user_id = reminder.bitrix_user_id
        active_structure = SalaryStructure.objects.filter(bitrix_user_id=bitrix_user_id).order_by('-effective_from').first()
        if not active_structure:
            return Response({'error': 'Active salary structure not found.'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = SalaryIncrementApprovalSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)

        new_basic = serializer.validated_data['new_basic']
        new_hra = serializer.validated_data['new_hra']
        new_allowances = serializer.validated_data['new_allowances']
        effective_date = serializer.validated_data.get('effective_date', reminder.anniversary_date)
        reason = serializer.validated_data['reason']

        old_net = active_structure.net_salary

        new_structure = SalaryStructure.objects.create(
            bitrix_user_id=bitrix_user_id,
            effective_from=effective_date,
            gross_salary=new_basic,
            pf_contribution=active_structure.pf_contribution,
            esi=active_structure.esi,
            labour_welfare_fund=active_structure.labour_welfare_fund,
            professional_tax=active_structure.professional_tax,
            other_deductions=active_structure.other_deductions
        )

        new_net = new_structure.net_salary
        increment_amount = new_net - old_net
        increment_pct = (increment_amount / old_net) * 100 if old_net > 0 else 0

        approval = serializer.save(
            bitrix_user_id=bitrix_user_id,
            reminder=reminder,
            old_net=old_net,
            new_net=new_net,
            increment_amount=increment_amount,
            increment_pct=increment_pct,
            approved_by=request.user
        )

        reminder.status = 'Actioned'
        reminder.actioned_by = request.user
        reminder.actioned_at = timezone.now()
        reminder.save()

        generate_increment_letter_pdf(approval, user=request.user)

        return Response({
            'message': 'Salary increment approved and new structure applied.',
            'approval': SalaryIncrementApprovalSerializer(approval).data
        }, status=status.HTTP_201_CREATED)


# New Endpoints
class SalaryExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can export salary sheets.")

        month_param = request.query_params.get('month')
        year_param = request.query_params.get('year')

        if not month_param or not year_param:
            return Response({'error': 'month and year parameters are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month = int(month_param)
            year = int(year_param)
        except ValueError:
            return Response({'error': 'Invalid month or year.'}, status=status.HTTP_400_BAD_REQUEST)

        # Columns
        headers = [
            "Sr. No.", "Name", "Designation", "Month days", "Worked days", 
            "Weekend", "CL", "Extra", "Payable Days", "Month Salary", 
            "Payable Salary", "Extra days working", "Fine/Advance", "Net Payable", 
            "Bank A/c No.", "Bank"
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Salary_{month}_{year}"
        ws.append(headers)

        # Apply sheet protection
        ws.protection.sheet = True
        ws.protection.password = 'mtlv_payroll'

        from openpyxl.styles import Protection
        locked_style = Protection(locked=True)
        unlocked_style = Protection(locked=False)

        # Check if salary slips already exist
        slips = SalarySlip.objects.filter(month=month, year=year)
        
        # Resolve employees for all slips
        user_map = {}
        for u in BitrixClient.get_all_users():
            user_map[str(u['id'])] = BitrixEmployeeMock(u)
            
        row_idx = 2
        from salary.models import EmployeeBankDetail

        if slips.exists():
            # Create a mapping of bitrix_user_id to slip for accurate employee matching
            slips_map = {str(slip.bitrix_user_id): slip for slip in slips}
            
            # Iterate through employees in consistent order to ensure correct mapping
            for emp in user_map.values():
                slip = slips_map.get(str(emp.bitrix_id))
                if not slip:
                    continue
                    
                emp_name = emp.name if emp else f"User {slip.bitrix_user_id}"
                designation = emp.designation if emp else ""
                
                detail = EmployeeBankDetail.objects.filter(bitrix_user_id=slip.bitrix_user_id).first()
                bank_acc = detail.bank_account_no if (detail and detail.bank_account_no) else (slip.bank_account_no or "")
                bank_nm = detail.bank_name if (detail and detail.bank_name) else (slip.bank_name or "")

                row_data = [
                    row_idx - 1,
                    emp_name,
                    designation,
                    float(slip.month_days),
                    float(slip.worked_days),
                    float(slip.weekend),
                    float(slip.cl),
                    float(slip.extra),
                    float(slip.payable_days),
                    float(slip.month_salary),
                    float(slip.payable_salary),
                    float(slip.extra_days_working),
                    float(slip.fine_advance),
                    float(slip.net_payable),
                    bank_acc,
                    bank_nm
                ]
                ws.append(row_data)
                
                # Lock columns 1, 2, 3, unlock others
                for col_idx in range(1, 17):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if col_idx in [1, 2, 3]:
                        cell.protection = locked_style
                    else:
                        cell.protection = unlocked_style
                row_idx += 1
        else:
            # Export with carry-forward logic: use latest prior imported salary if available
            import calendar
            from exit_formality.models import ExitRequest
            month_days_val = calendar.monthrange(year, month)[1]
            
            # Use the same user_map for consistent employee ordering
            employees = []
            for emp in user_map.values():
                if emp.status != 'Exited':
                    employees.append(emp)
                else:
                    # Include exited employees if they have a pending/active exit request
                    has_pending_exit = ExitRequest.objects.filter(
                        bitrix_user_id=emp.bitrix_id
                    ).exclude(status__in=['CANCELLED', 'FULLY_EXITED']).exists()
                    if has_pending_exit:
                        employees.append(emp)
            
            for emp in employees:
                # Try to get latest prior imported salary for carry-forward
                prior_slip = get_latest_prior_slip(emp.bitrix_id, month, year)
                
                # Get salary structure for new employees (no prior imports)
                struct = None
                if not prior_slip:
                    struct = SalaryStructure.objects.filter(
                        bitrix_user_id=emp.bitrix_id,
                        effective_from__lte=datetime.date(year, month, 28)
                    ).order_by('-effective_from').first()
                    if not struct:
                        struct = SalaryStructure.objects.filter(
                            bitrix_user_id=emp.bitrix_id
                        ).order_by('-effective_from').first()

                detail = EmployeeBankDetail.objects.filter(bitrix_user_id=emp.bitrix_id).first()
                bank_acc = (detail.bank_account_no if detail and detail.bank_account_no else None) or \
                           (prior_slip.bank_account_no if prior_slip else None) or \
                           getattr(emp, 'bank_account', '') or ""
                bank_nm = (detail.bank_name if detail and detail.bank_name else None) or \
                          (prior_slip.bank_name if prior_slip else None) or ""

                # Use carry-forward data if available, otherwise blank with structure gross
                if prior_slip:
                    row_data = [
                        row_idx - 1,
                        emp.name,
                        emp.designation,
                        float(month_days_val),          # always use actual month days
                        float(prior_slip.worked_days),
                        float(prior_slip.weekend),
                        float(prior_slip.cl),
                        float(prior_slip.extra),
                        float(prior_slip.payable_days),
                        float(prior_slip.month_salary),
                        float(prior_slip.payable_salary),
                        float(prior_slip.extra_days_working),
                        float(prior_slip.fine_advance),
                        float(prior_slip.net_payable),
                        bank_acc,
                        bank_nm
                    ]
                else:
                    # New employee — blank row with only gross from structure
                    row_data = [
                        row_idx - 1,
                        emp.name,
                        emp.designation,
                        float(month_days_val),
                        0.0,  # Worked days
                        0.0,  # Weekend
                        0.0,  # CL
                        0.0,  # Extra
                        0.0,  # Payable Days
                        float(struct.gross_salary) if struct else 0.0,
                        0.0,  # Payable Salary
                        0.0,  # Extra days working
                        0.0,  # Fine/Advance
                        0.0,  # Net Payable
                        bank_acc,
                        bank_nm
                ]
                ws.append(row_data)

                for col_idx in range(1, 17):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if col_idx in [1, 2, 3]:
                        cell.protection = locked_style
                    else:
                        cell.protection = unlocked_style
                row_idx += 1

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="salary_sheet_{month}_{year}.xlsx"'
        wb.save(response)
        return response


class SalaryImportView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can import salary sheets.")

        file_obj = request.FILES.get('file')
        month_str = request.data.get('month')
        year_str = request.data.get('year')

        if not file_obj or not month_str or not year_str:
            return Response({'error': 'file, month, and year are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month = int(month_str)
            year = int(year_str)
            if not (1 <= month <= 12):
                raise ValueError()
        except ValueError:
            return Response({'error': 'Invalid month or year.'}, status=status.HTTP_400_BAD_REQUEST)

        # 1. Create batch
        batch = SalaryImportBatch.objects.create(
            month=month,
            year=year,
            file_name=file_obj.name,
            uploaded_by=request.user,
            status='processing'
        )

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
        except Exception as e:
            batch.status = 'failed'
            batch.save()
            return Response({'error': f'Failed to parse Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # Parsers
        def parse_decimal(val, col_name):
            if val is None or str(val).strip() == "":
                return Decimal("0.00")
            try:
                cleaned = str(val).strip().replace("$", "").replace(",", "")
                d = Decimal(cleaned)
                if d < 0:
                    raise ValueError()
                return d
            except Exception:
                raise ValueError(f"Invalid value in column {col_name}")

        headers = [
            "Sr. No.", "Name", "Designation", "Month days", "Worked days", 
            "Weekend", "CL", "Extra", "Payable Days", "Month Salary", 
            "Payable Salary", "Extra days working", "Fine/Advance", "Net Payable", 
            "Bank A/c No.", "Bank"
        ]

        total_records = 0
        success_count = 0
        failed_count = 0
        failed_rows_data = []

        try:
            with transaction.atomic():
                for r_idx in range(2, ws.max_row + 1):
                    name_val = ws.cell(row=r_idx, column=2).value
                    if name_val is None or str(name_val).strip() == "":
                        continue

                    total_records += 1
                    row_vals = [ws.cell(row=r_idx, column=c).value for c in range(1, 17)]
                    name_str = str(name_val).strip()

                    sid = transaction.savepoint()
                    try:
                        # Validate Employee
                        bitrix_users = BitrixClient.get_all_users()
                        employee = None
                        for u in bitrix_users:
                            mock_emp = BitrixEmployeeMock(u)
                            if mock_emp.name.lower().strip() == name_str.lower():
                                employee = mock_emp
                                break
                        if not employee:
                            raise ValueError("Employee not found in Bitrix24")
                        if employee.status == 'Exited':
                            from exit_formality.models import ExitRequest
                            has_pending_exit = ExitRequest.objects.filter(
                                bitrix_user_id=employee.bitrix_id
                            ).exclude(status__in=['CANCELLED', 'FULLY_EXITED']).exists()
                            if not has_pending_exit:
                                raise ValueError("Employee is exited and has no pending exit request in the system")

                        # Parse data
                        month_days = parse_decimal(ws.cell(row=r_idx, column=4).value, "Month days")
                        worked_days = parse_decimal(ws.cell(row=r_idx, column=5).value, "Worked days")
                        weekend = parse_decimal(ws.cell(row=r_idx, column=6).value, "Weekend")
                        cl = parse_decimal(ws.cell(row=r_idx, column=7).value, "CL")
                        extra = parse_decimal(ws.cell(row=r_idx, column=8).value, "Extra")
                        payable_days = parse_decimal(ws.cell(row=r_idx, column=9).value, "Payable Days")
                        month_salary = parse_decimal(ws.cell(row=r_idx, column=10).value, "Month Salary")
                        payable_salary = parse_decimal(ws.cell(row=r_idx, column=11).value, "Payable Salary")
                        extra_days_working = parse_decimal(ws.cell(row=r_idx, column=12).value, "Extra days working")
                        fine_advance = parse_decimal(ws.cell(row=r_idx, column=13).value, "Fine/Advance")
                        net_payable = parse_decimal(ws.cell(row=r_idx, column=14).value, "Net Payable")
                        
                        bank_account_val = ws.cell(row=r_idx, column=15).value
                        if bank_account_val is not None:
                            if isinstance(bank_account_val, float):
                                bank_account_no = str(int(bank_account_val)).strip()
                            else:
                                bank_account_no = str(bank_account_val).strip()
                        else:
                            bank_account_no = ""
                            
                        bank_name = str(ws.cell(row=r_idx, column=16).value or "").strip()

                        from salary.models import EmployeeBankDetail
                        # Prefill from EmployeeBankDetail if not provided in the excel row
                        if not bank_account_no or not bank_name:
                            detail = EmployeeBankDetail.objects.filter(bitrix_user_id=employee.bitrix_id).first()
                            if detail:
                                if not bank_account_no:
                                    bank_account_no = detail.bank_account_no or ""
                                if not bank_name:
                                    bank_name = detail.bank_name or ""

                        # If both are provided, save or update EmployeeBankDetail
                        if bank_account_no and bank_name:
                            EmployeeBankDetail.objects.update_or_create(
                                bitrix_user_id=employee.bitrix_id,
                                defaults={
                                    'bank_account_no': bank_account_no,
                                    'bank_name': bank_name
                                }
                            )

                        # Save SalarySlip
                        slip = SalarySlip.objects.filter(bitrix_user_id=employee.bitrix_id, month=month, year=year).first()
                        if slip:
                            slip.month_days = month_days
                            slip.worked_days = worked_days
                            slip.weekend = weekend
                            slip.cl = cl
                            slip.extra = extra
                            slip.payable_days = payable_days
                            slip.month_salary = month_salary
                            slip.payable_salary = payable_salary
                            slip.extra_days_working = extra_days_working
                            slip.fine_advance = fine_advance
                            slip.net_payable = net_payable
                            slip.bank_account_no = bank_account_no
                            slip.bank_name = bank_name
                            slip.status = 'draft' # Re-review needed
                            slip.uploaded_batch = batch
                            slip._skip_recalculation = True
                            slip.save()
                        else:
                            slip = SalarySlip(
                                bitrix_user_id=employee.bitrix_id,
                                month=month,
                                year=year,
                                month_days=month_days,
                                worked_days=worked_days,
                                weekend=weekend,
                                cl=cl,
                                extra=extra,
                                payable_days=payable_days,
                                month_salary=month_salary,
                                payable_salary=payable_salary,
                                extra_days_working=extra_days_working,
                                fine_advance=fine_advance,
                                net_payable=net_payable,
                                bank_account_no=bank_account_no,
                                bank_name=bank_name,
                                status='draft',
                                uploaded_batch=batch
                            )
                            slip._skip_recalculation = True
                            slip.save()
                        
                        generate_payslip_pdf(slip)
                        transaction.savepoint_commit(sid)
                        success_count += 1
                    except Exception as e:
                        transaction.savepoint_rollback(sid)
                        failed_count += 1
                        failed_rows_data.append((row_vals, str(e)))

                if failed_count == total_records and total_records > 0:
                    raise Exception("All rows failed validation")
        except Exception as batch_err:
            # Batch fully failed rollback occurred
            batch.status = 'failed'
            batch.total_records = total_records
            batch.success_count = 0
            batch.failed_count = total_records
            
            error_report_url = self.generate_error_xlsx(batch.id, headers, failed_rows_data)
            batch.error_report_path = error_report_url
            batch.save()
            return Response({
                "total": total_records,
                "success": 0,
                "failed": total_records,
                "error_report_url": error_report_url,
                "detail": str(batch_err)
            }, status=status.HTTP_400_BAD_REQUEST)

        # Batch committed
        if failed_count > 0:
            batch.status = 'partial'
            error_report_url = self.generate_error_xlsx(batch.id, headers, failed_rows_data)
            batch.error_report_path = error_report_url
            
            # Trigger 21: Salary Import Completed with Errors
            try:
                from notifications.models import Notification
                from django.contrib.auth import get_user_model
                from django.core.mail import send_mail
                
                User = get_user_model()
                admins = User.objects.filter(role__code='ADMIN')
                hrs = User.objects.filter(role__code='HR')
                
                msg = f"Salary import completed. {success_count} rows imported successfully. {failed_count} rows failed. Download error report to review."
                
                for hr in hrs:
                    Notification.objects.create(
                        recipient=hr,
                        notif_type='WARNING',
                        message=msg,
                        link="/salary/import/result/"
                    )
                    if hr.email:
                        error_full_url = f"{settings.FRONTEND_URL}{error_report_url}" if hasattr(settings, 'FRONTEND_URL') else f"http://localhost:8000{error_report_url}"
                        send_mail(
                            subject="Salary Import Completed with Errors",
                            message=f"{msg}\n\nDownload the error report here: {error_full_url}",
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            recipient_list=[hr.email],
                            fail_silently=True
                        )
                for admin in admins:
                    Notification.objects.create(
                        recipient=admin,
                        notif_type='WARNING',
                        message=msg,
                        link="/salary/import/result/"
                    )
            except Exception:
                pass
        else:
            batch.status = 'success'
            error_report_url = None

        batch.total_records = total_records
        batch.success_count = success_count
        batch.failed_count = failed_count
        batch.save()

        return Response({
            "total": total_records,
            "success": success_count,
            "failed": failed_count,
            "error_report_url": error_report_url
        }, status=status.HTTP_200_OK)

    def generate_error_xlsx(self, batch_id, headers, failed_rows_data):
        error_report_dir = os.path.join(settings.MEDIA_ROOT, "error_reports")
        os.makedirs(error_report_dir, exist_ok=True)
        report_filename = f"error_report_{batch_id}_{int(datetime.datetime.now().timestamp())}.xlsx"
        report_path = os.path.join(error_report_dir, report_filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Errors"
        ws.append(headers + ["Error Reason"])

        for row_vals, reason in failed_rows_data:
            ws.append(row_vals + [reason])

        wb.save(report_path)
        return f"{settings.MEDIA_URL}error_reports/{report_filename}"


class SalaryPublishView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can publish slips.")

        month_str = request.data.get('month')
        year_str = request.data.get('year')

        if not month_str or not year_str:
            return Response({'error': 'month and year are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month = int(month_str)
            year = int(year_str)
        except ValueError:
            return Response({'error': 'Invalid month or year.'}, status=status.HTTP_400_BAD_REQUEST)

        slips = SalarySlip.objects.filter(month=month, year=year, status='draft')
        count = slips.count()
        for slip in slips:
            slip.status = 'published'
            slip.save()
            generate_payslip_pdf(slip)

        # Trigger 20 notification
        import calendar
        try:
            month_name = calendar.month_name[month]
            month_year = f"{month_name} {year}"
            month_year_slug = f"{month_name.lower()}-{year}"
            
            from notifications.models import Notification
            from django.contrib.auth import get_user_model
            User = get_user_model()
            admins_and_hrs = User.objects.filter(role__code__in=['ADMIN', 'HR'])
            msg = f"Salary slips generated for {month_year}. {count} slips ready for download."
            for user in admins_and_hrs:
                Notification.objects.create(
                    recipient=user,
                    notif_type='SUCCESS',
                    message=msg,
                    link=f"/salary/slips/{month_year_slug}/"
                )
        except Exception:
            pass

        return Response({'message': f'Successfully published {count} salary slips.'}, status=status.HTTP_200_OK)


class SalaryEditView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can edit salary slips.")

        slip = get_object_or_404(SalarySlip, pk=pk)

        # Exclude read-only calculations from direct parsing
        mutable_fields = [
            'location', 'month_days', 'worked_days', 'weekend', 'cl', 'extra',
            'month_salary', 'extra_days_working', 'fine_advance',
            'bank_account_no', 'bank_name',
            'leaves_available', 'working_days', 'extra_days',
            'gross_salary', 'pf_contribution', 'esi', 'labour_welfare_fund',
            'professional_tax', 'other_deductions', 'net_credited_amount',
            'payment_status', 'payment_date', 'transaction_ref'
        ]

        for field in mutable_fields:
            if field in request.data:
                val = request.data[field]
                if field in ['payment_status', 'transaction_ref', 'location', 'bank_account_no', 'bank_name']:
                    setattr(slip, field, val)
                elif field == 'payment_date':
                    setattr(slip, field, val if val else None)
                else:
                    setattr(slip, field, Decimal(str(val)) if val else Decimal('0.00'))

        slip.status = 'draft' # Re-review needed after edit
        slip.save()

        # Update or create persistent employee bank details if both are present in the slip
        if slip.bank_account_no and slip.bank_name:
            from salary.models import EmployeeBankDetail
            EmployeeBankDetail.objects.update_or_create(
                bitrix_user_id=slip.bitrix_user_id,
                defaults={
                    'bank_account_no': slip.bank_account_no,
                    'bank_name': slip.bank_name
                }
            )

        generate_payslip_pdf(slip)

        serializer = SalarySlipSerializer(slip)
        return Response(serializer.data, status=status.HTTP_200_OK)


class SalaryHistoryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        role = check_role(request.user)
        
        # Scoped logic
        if role == 'employee':
            # Force self
            user_data = next((u for u in BitrixClient.get_all_users() if u.get('email') == request.user.email), None)
            if not user_data:
                return Response({'error': 'Employee profile not found in Bitrix24.'}, status=status.HTTP_404_NOT_FOUND)
            employee = BitrixEmployeeMock(user_data)
            slips = SalarySlip.objects.filter(bitrix_user_id=employee.bitrix_id, status='published').order_by('-year', '-month')
        else:
            # Admin / HR
            employee_id = request.query_params.get('employee_id') or kwargs.get('employee_id') or request.query_params.get('bitrix_user_id')
            if employee_id:
                slips = SalarySlip.objects.filter(bitrix_user_id=str(employee_id)).order_by('-year', '-month')
            else:
                # Paginate all employees slips
                slips = SalarySlip.objects.all().order_by('-year', '-month')

        # Payment status filtering
        payment_status = request.query_params.get('payment_status')
        if payment_status:
            slips = slips.filter(payment_status=payment_status)

        # Date range filtering
        from_param = request.query_params.get('from') # format: YYYY-MM
        to_param = request.query_params.get('to')     # format: YYYY-MM

        if from_param:
            try:
                fY, fM = map(int, from_param.split('-'))
                slips = slips.filter(
                    models.Q(year__gt=fY) | models.Q(year=fY, month__gte=fM)
                )
            except ValueError:
                pass

        if to_param:
            try:
                tY, tM = map(int, to_param.split('-'))
                slips = slips.filter(
                    models.Q(year__lt=tY) | models.Q(year=tY, month__lte=tM)
                )
            except ValueError:
                pass

        # Pagination
        paginator = StandardResultsSetPagination()
        page = paginator.paginate_queryset(slips, request, view=self)
        
        bitrix_users = {str(u['id']): BitrixEmployeeMock(u) for u in BitrixClient.get_all_users()}
        data = []
        target_list = page if page is not None else slips
        for s in target_list:
            emp = bitrix_users.get(str(s.bitrix_user_id))
            emp_id = emp.id if emp else 0
            emp_name = emp.name if emp else f"User {s.bitrix_user_id}"
            data.append({
                'id': s.id,
                'employee_id': emp_id,
                'employee_name': emp_name,
                'month': s.month,
                'year': s.year,
                'location': s.location,
                'leaves_available': str(s.leaves_available),
                'working_days': str(s.working_days),
                'extra_days': str(s.extra_days),
                'gross_salary': str(s.gross_salary),
                'pf_contribution': str(s.pf_contribution),
                'esi': str(s.esi),
                'labour_welfare_fund': str(s.labour_welfare_fund),
                'professional_tax': str(s.professional_tax),
                'other_deductions': str(s.other_deductions),
                'total_deductions': str(s.total_deductions),
                'net_salary': str(s.net_salary),
                'net_credited_amount': str(s.net_credited_amount),
                'month_days': str(s.month_days),
                'worked_days': str(s.worked_days),
                'weekend': str(s.weekend),
                'cl': str(s.cl),
                'extra': str(s.extra),
                'payable_days': str(s.payable_days),
                'month_salary': str(s.month_salary),
                'payable_salary': str(s.payable_salary),
                'extra_days_working': str(s.extra_days_working),
                'fine_advance': str(s.fine_advance),
                'net_payable': str(s.net_payable),
                'bank_account_no': s.bank_account_no or "",
                'bank_name': s.bank_name or "",
                'payment_status': s.payment_status,
                'payment_date': str(s.payment_date) if s.payment_date else '',
                'transaction_ref': s.transaction_ref or '',
                'status': s.status
            })

        if page is not None:
            return paginator.get_paginated_response(data)
        return Response(data)


class SalarySlipDownloadView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        role = check_role(request.user)

        employee_id = request.query_params.get('employee_id')
        download_type = request.query_params.get('type') # single | last3 | last4 | range | bulk_month | selected

        if role == 'employee':
            user_data = next((u for u in BitrixClient.get_all_users() if u.get('email') == request.user.email), None)
            if not user_data:
                return Response({'error': 'Employee profile not found.'}, status=status.HTTP_404_NOT_FOUND)
            employee = BitrixEmployeeMock(user_data)
            employee_id = employee.bitrix_id
            # Employees can only download published slips
            slips_qs = SalarySlip.objects.filter(bitrix_user_id=employee_id, status='published')
        else:
            # Admin/HR
            if download_type in ['bulk_month', 'selected'] and not employee_id:
                slips_qs = SalarySlip.objects.all()
            else:
                if not employee_id:
                    return Response({'error': 'employee_id parameter is required for Admin/HR.'}, status=status.HTTP_400_BAD_REQUEST)
                slips_qs = SalarySlip.objects.filter(bitrix_user_id=str(employee_id))

        # Filters
        slips = []
        if download_type == 'single':
            month_str = request.query_params.get('month')
            year_str = request.query_params.get('year')
            if not month_str or not year_str:
                return Response({'error': 'month and year are required.'}, status=status.HTTP_400_BAD_REQUEST)
            slips = slips_qs.filter(month=int(month_str), year=int(year_str))
        elif download_type == 'last3':
            slips = slips_qs.order_by('-year', '-month')[:3]
        elif download_type == 'last4':
            slips = slips_qs.order_by('-year', '-month')[:4]
        elif download_type == 'range':
            from_month = int(request.query_params.get('from_month', 1))
            from_year = int(request.query_params.get('from_year', 2000))
            to_month = int(request.query_params.get('to_month', 12))
            to_year = int(request.query_params.get('to_year', 2100))
            
            # Filters BETWEEN from and to inclusive
            slips = slips_qs.filter(
                models.Q(year__gt=from_year) | models.Q(year=from_year, month__gte=from_month)
            ).filter(
                models.Q(year__lt=to_year) | models.Q(year=to_year, month__lte=to_month)
            ).order_by('year', 'month')
        elif download_type == 'bulk_month':
            if role not in ['admin', 'hr']:
                raise PermissionDenied("Only Admin/HR can bulk download slips.")
            month_str = request.query_params.get('month')
            year_str = request.query_params.get('year')
            if not month_str or not year_str:
                return Response({'error': 'month and year are required.'}, status=status.HTTP_400_BAD_REQUEST)
            slips = slips_qs.filter(month=int(month_str), year=int(year_str))
        elif download_type == 'selected':
            slip_ids_str = request.query_params.get('slip_ids')
            if not slip_ids_str:
                return Response({'error': 'slip_ids parameter is required for type=selected.'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                ids = [int(x) for x in slip_ids_str.split(',')]
                slips = slips_qs.filter(id__in=ids)
            except ValueError:
                return Response({'error': 'Invalid slip_ids parameter.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Invalid download type.'}, status=status.HTTP_400_BAD_REQUEST)

        slips = list(slips)
        if not slips:
            return Response({'error': 'No salary slips found matching parameters.'}, status=status.HTTP_404_NOT_FOUND)

        if len(slips) == 1:
            slip = slips[0]
            # Ensure PDF exists
            generate_payslip_pdf(slip)
            
            # Serve single PDF
            response = HttpResponse(slip.pdf_file.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="payslip_{slip.employee.emp_id}_{slip.month}_{slip.year}.pdf"'
            return response
        else:
            # Serve ZIP
            zip_type = 'bulk' if download_type == 'bulk_month' else 'employee'
            zip_data = generate_payslips_zip(slips, zip_type)
            
            response = HttpResponse(zip_data, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="payslips_{int(datetime.datetime.now().timestamp())}.zip"'
            return response


class SalaryImportBatchesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can view import batches.")

        batches = SalaryImportBatch.objects.all().order_by('-uploaded_at')
        paginator = StandardResultsSetPagination()
        page = paginator.paginate_queryset(batches, request, view=self)
        serializer = SalaryImportBatchSerializer(page if page is not None else batches, many=True)
        
        if page is not None:
            return paginator.get_paginated_response(serializer.data)
        return Response(serializer.data)


class SalaryEmployeeSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, employee_id):
        role = check_role(request.user)
        if role == 'employee':
            # Employees can only view their own summary
            user_data = next((u for u in BitrixClient.get_all_users() if u.get('email') == request.user.email), None)
            if not user_data or str(user_data.get('id')) != str(employee_id):
                raise PermissionDenied("You can only view your own salary summary.")

        slips = SalarySlip.objects.filter(bitrix_user_id=str(employee_id))

        # Payment status filtering
        payment_status = request.query_params.get('payment_status')
        if payment_status:
            slips = slips.filter(payment_status=payment_status)

        # Date range filtering
        from_param = request.query_params.get('from') # format: YYYY-MM
        to_param = request.query_params.get('to')     # format: YYYY-MM

        if from_param:
            try:
                fY, fM = map(int, from_param.split('-'))
                slips = slips.filter(
                    models.Q(year__gt=fY) | models.Q(year=fY, month__gte=fM)
                )
            except ValueError:
                pass

        if to_param:
            try:
                tY, tM = map(int, to_param.split('-'))
                slips = slips.filter(
                    models.Q(year__lt=tY) | models.Q(year=tY, month__lte=tM)
                )
            except ValueError:
                pass
        
        total_credited = sum((slip.net_payable or Decimal('0.00') for slip in slips), Decimal('0.00'))
        total_deductions = sum((slip.fine_advance or Decimal('0.00') for slip in slips), Decimal('0.00'))
        total_payslips = slips.count()
        last_paid_slip = slips.filter(payment_status='paid', payment_date__isnull=False).order_by('-payment_date').first()
        last_payment_date = last_paid_slip.payment_date.strftime('%Y-%m-%d') if last_paid_slip else '-'

        return Response({
            'total_salary_credited': str(total_credited),
            'total_deductions': str(total_deductions),
            'total_payslips': total_payslips,
            'last_payment_date': last_payment_date
        })


class SalaryEmployeeHistoryExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, employee_id):
        role = check_role(request.user)
        if role == 'employee':
            user_data = next((u for u in BitrixClient.get_all_users() if u.get('email') == request.user.email), None)
            if not user_data or str(user_data.get('id')) != str(employee_id):
                raise PermissionDenied("You can only export your own salary history.")
        else:
            user_data = BitrixClient.get_user_detail(employee_id)

        if not user_data:
            return Response({'error': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        employee = BitrixEmployeeMock(user_data)
        slips = SalarySlip.objects.filter(bitrix_user_id=employee.bitrix_id).order_by('-year', '-month')

        from_param = request.query_params.get('from')
        to_param = request.query_params.get('to')
        payment_status = request.query_params.get('payment_status')

        if from_param:
            try:
                fY, fM = map(int, from_param.split('-'))
                slips = slips.filter(models.Q(year__gt=fY) | models.Q(year=fY, month__gte=fM))
            except ValueError:
                pass
        if to_param:
            try:
                tY, tM = map(int, to_param.split('-'))
                slips = slips.filter(models.Q(year__lt=tY) | models.Q(year=tY, month__lte=tM))
            except ValueError:
                pass
        if payment_status:
            slips = slips.filter(payment_status=payment_status)

        headers = [
            "Month/Year", "Location", "Month days", "Worked days", "Weekend", 
            "CL", "Extra", "Payable Days", "Month Salary", "Payable Salary", 
            "Extra days working", "Fine/Advance", "Net Payable", "Bank A/c No.", "Bank",
            "Payment Status", "Payment Date", "Transaction Ref"
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Salary History"
        ws.append(headers)

        month_names = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

        for s in slips:
            ws.append([
                f"{month_names[s.month]} {s.year}",
                s.location,
                float(s.month_days),
                float(s.worked_days),
                float(s.weekend),
                float(s.cl),
                float(s.extra),
                float(s.payable_days),
                float(s.month_salary),
                float(s.payable_salary),
                float(s.extra_days_working),
                float(s.fine_advance),
                float(s.net_payable),
                s.bank_account_no or "",
                s.bank_name or "",
                s.payment_status.capitalize(),
                s.payment_date.strftime('%Y-%m-%d') if s.payment_date else '',
                s.transaction_ref or ''
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="salary_history_{employee.emp_id}.xlsx"'
        wb.save(response)
        return response
