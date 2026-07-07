import datetime
import os
import openpyxl
from decimal import Decimal
from django.db import models, transaction
from django.core.files.base import ContentFile
from django.conf import settings
from django.http import HttpResponse, Http404
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import viewsets, status, serializers
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.pagination import PageNumberPagination

from common.bitrix_client import BitrixClient, BitrixEmployeeMock

from .models import (
    DismissedSalaryStructure, DismissedSalarySlip, DismissedSalaryImportBatch,
    DismissedEmployeeBankDetail
)
from .dismissed_serializers import (
    DismissedSalaryStructureSerializer, DismissedSalarySlipSerializer,
    DismissedSalaryImportBatchSerializer
)
from .services import (
    generate_dismissed_payslip_pdf, generate_dismissed_payslips_zip,
    generate_dismissed_payslip_pdf_bytes, get_latest_prior_dismissed_slip,
    calculate_carry_forward_dismissed_slip, num_to_words
)
from .views import check_role, StandardResultsSetPagination

# Legacy viewsets for backward compatibility with UI

class DismissedSalaryStructureViewSet(viewsets.ModelViewSet):
    serializer_class = DismissedSalaryStructureSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = DismissedSalaryStructure.objects.all().order_by('-effective_from')
        employee_id = self.request.query_params.get('employee_id')
        if employee_id:
            queryset = queryset.filter(bitrix_user_id=employee_id)
        return queryset

    def perform_create(self, serializer):
        # Admin check
        if check_role(self.request.user) != 'admin':
            raise PermissionDenied("Only Admin can manage salary structures.")
        serializer.save()


class SalaryIncrementViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'approve':
            return SalaryIncrementApprovalSerializer
        return SalaryIncrementReminderSerializer

    def get_queryset(self):
        if self.action in ('list_approvals', 'retrieve_approval'):
            return SalaryIncrementApproval.objects.all().order_by('-approved_at')
        return SalaryIncrementReminder.objects.all().order_by('-anniversary_date')

    @action(detail=False, methods=['GET'], url_path='approvals')
    def list_approvals(self, request):
        queryset = SalaryIncrementApproval.objects.all().order_by('-approved_at')
        serializer = SalaryIncrementApprovalSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['POST'], url_path='approve')
    @transaction.atomic
    def approve(self, request, pk=None):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can approve increments.")

        reminder = get_object_or_404(SalaryIncrementReminder, pk=pk)
        if reminder.status == 'Actioned':
            return Response({'error': 'This increment reminder has already been actioned.'}, status=status.HTTP_400_BAD_REQUEST)

        bitrix_user_id = reminder.bitrix_user_id
        active_structure = DismissedSalaryStructure.objects.filter(bitrix_user_id=bitrix_user_id).order_by('-effective_from').first()
        if not active_structure:
            return Response({'error': 'Active salary structure not found.'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = SalaryIncrementApprovalSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)

        new_basic = serializer.validated_data['new_basic']
        new_hra = serializer.validated_data['new_hra']
        new_allowances = serializer.validated_data['new_allowances']
        effective_date = serializer.validated_data.get('effective_date', reminder.anniversary_date)
        reason = serializer.validated_data['reason']

        old_net = active_structure.net_salary

        new_structure = DismissedSalaryStructure.objects.create(
            bitrix_user_id=bitrix_user_id,
            effective_from=effective_date,
            gross_salary=new_basic,
            pf_contribution=active_structure.pf_contribution,
            esi=active_structure.esi,
            labour_welfare_fund=active_structure.labour_welfare_fund,
            professional_tax=active_structure.professional_tax,
            other_deductions=active_structure.other_deductions
        )

        new_net = new_structure.net_salary
        increment_amount = new_net - old_net
        increment_pct = (increment_amount / old_net) * 100 if old_net > 0 else 0

        approval = serializer.save(
            bitrix_user_id=bitrix_user_id,
            reminder=reminder,
            old_net=old_net,
            new_net=new_net,
            increment_amount=increment_amount,
            increment_pct=increment_pct,
            approved_by=request.user
        )

        reminder.status = 'Actioned'
        reminder.actioned_by = request.user
        reminder.actioned_at = timezone.now()
        reminder.save()

        generate_increment_letter_pdf(approval, user=request.user)

        return Response({
            'message': 'Salary increment approved and new structure applied.',
            'approval': SalaryIncrementApprovalSerializer(approval).data
        }, status=status.HTTP_201_CREATED)


# New Endpoints
class DismissedSalaryExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can export salary sheets.")

        month_param = request.query_params.get('month')
        year_param = request.query_params.get('year')

        if not month_param or not year_param:
            return Response({'error': 'month and year parameters are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month = int(month_param)
            year = int(year_param)
        except ValueError:
            return Response({'error': 'Invalid month or year.'}, status=status.HTTP_400_BAD_REQUEST)

        # Columns
        headers = [
            "Sr. No.", "Name", "Designation", "Month days", "Worked days", 
            "Weekend", "CL", "Extra", "Payable Days", "Month Salary", 
            "Payable Salary", "Extra days working", "Fine/Advance", "Net Payable", 
            "Bank A/c No.", "Bank"
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Salary_{month}_{year}"
        ws.append(headers)

        # Apply sheet protection
        ws.protection.sheet = True
        ws.protection.password = 'mtlv_payroll'

        from openpyxl.styles import Protection
        locked_style = Protection(locked=True)
        unlocked_style = Protection(locked=False)

        slips = DismissedSalarySlip.objects.filter(month=month, year=year)
        slips_map = {str(slip.bitrix_user_id): slip for slip in slips}
        
        # Resolve employees for all slips
        user_map = {}
        for u in BitrixClient.get_all_users():
            user_map[str(u['id'])] = BitrixEmployeeMock(u)
            
        row_idx = 2
        from salary.models import DismissedEmployeeBankDetail
        import calendar
        from exit_formality.models import ExitRequest
        month_days_val = calendar.monthrange(year, month)[1]

        # Use the same user_map for consistent employee ordering
        employees = []
        for emp in user_map.values():
            if emp.status == 'Exited':
                employees.append(emp)
        
        for emp in employees:
            slip = slips_map.get(str(emp.bitrix_id))
            
            if slip:
                emp_name = emp.name if emp else f"User {slip.bitrix_user_id}"
                designation = emp.designation if emp else ""
                
                detail = DismissedEmployeeBankDetail.objects.filter(bitrix_user_id=slip.bitrix_user_id).first()
                bank_acc = detail.bank_account_no if (detail and detail.bank_account_no) else (slip.bank_account_no or "")
                bank_nm = detail.bank_name if (detail and detail.bank_name) else (slip.bank_name or "")

                row_data = [
                    row_idx - 1,
                    emp_name,
                    designation,
                    float(slip.month_days),
                    float(slip.worked_days),
                    float(slip.weekend),
                    float(slip.cl),
                    float(slip.extra),
                    float(slip.payable_days),
                    float(slip.month_salary),
                    float(slip.payable_salary),
                    float(slip.extra_days_working),
                    float(slip.fine_advance),
                    float(slip.net_payable),
                    bank_acc,
                    bank_nm
                ]
                ws.append(row_data)
                
                # Lock columns 1, 2, 3, unlock others
                for col_idx in range(1, 17):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if col_idx in [1, 2, 3]:
                        cell.protection = locked_style
                    else:
                        cell.protection = unlocked_style
                row_idx += 1
            else:
                prior_slip = get_latest_prior_dismissed_slip(emp.bitrix_id, month, year)
                
                struct = None
                if not prior_slip:
                    import datetime
                    struct = DismissedSalaryStructure.objects.filter(
                        bitrix_user_id=emp.bitrix_id,
                        effective_from__lte=datetime.date(year, month, 28)
                    ).order_by('-effective_from').first()
                    if not struct:
                        struct = DismissedSalaryStructure.objects.filter(
                            bitrix_user_id=emp.bitrix_id
                        ).order_by('-effective_from').first()

                detail = DismissedEmployeeBankDetail.objects.filter(bitrix_user_id=emp.bitrix_id).first()
                bank_acc = (detail.bank_account_no if detail and detail.bank_account_no else None) or \
                           (prior_slip.bank_account_no if prior_slip else None) or \
                           getattr(emp, 'bank_account', '') or ""
                bank_nm = (detail.bank_name if detail and detail.bank_name else None) or \
                          (prior_slip.bank_name if prior_slip else None) or ""

                # Use carry-forward data if available, otherwise blank with structure gross
                if prior_slip:
                    temp_slip = calculate_carry_forward_dismissed_slip(emp.bitrix_id, month, year, prior_slip)
                    row_data = [
                        row_idx - 1,
                        emp.name,
                        emp.designation,
                        float(temp_slip.month_days),
                        float(temp_slip.worked_days),
                        float(temp_slip.weekend),
                        float(temp_slip.cl),
                        float(temp_slip.extra),
                        float(temp_slip.payable_days),
                        float(temp_slip.month_salary),
                        float(temp_slip.payable_salary),
                        float(temp_slip.extra_days_working),
                        float(temp_slip.fine_advance),
                        float(temp_slip.net_payable),
                        bank_acc,
                        bank_nm
                    ]
                else:
                    # New employee — blank row with only gross from structure
                    row_data = [
                        row_idx - 1,
                        emp.name,
                        emp.designation,
                        float(month_days_val),
                        0.0,  # Worked days
                        0.0,  # Weekend
                        0.0,  # CL
                        0.0,  # Extra
                        0.0,  # Payable Days
                        float(struct.gross_salary) if struct else 0.0,
                        0.0,  # Payable Salary
                        0.0,  # Extra days working
                        0.0,  # Fine/Advance
                        0.0,  # Net Payable
                        bank_acc,
                        bank_nm
                ]
                ws.append(row_data)

                for col_idx in range(1, 17):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if col_idx in [1, 2, 3]:
                        cell.protection = locked_style
                    else:
                        cell.protection = unlocked_style
                row_idx += 1

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="salary_sheet_{month}_{year}.xlsx"'
        wb.save(response)
        return response


class DismissedSalaryImportView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can import salary sheets.")

        file_obj = request.FILES.get('file')
        month_str = request.data.get('month')
        year_str = request.data.get('year')

        if not file_obj or not month_str or not year_str:
            return Response({'error': 'file, month, and year are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month = int(month_str)
            year = int(year_str)
            if not (1 <= month <= 12):
                raise ValueError()
        except ValueError:
            return Response({'error': 'Invalid month or year.'}, status=status.HTTP_400_BAD_REQUEST)

        # 1. Create batch
        batch = DismissedSalaryImportBatch.objects.create(
            month=month,
            year=year,
            file_name=file_obj.name,
            uploaded_by=request.user,
            status='processing'
        )

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
        except Exception as e:
            batch.status = 'failed'
            batch.save()
            return Response({'error': f'Failed to parse Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # Parsers
        def parse_decimal(val, col_name):
            if val is None or str(val).strip() == "":
                return Decimal("0.00")
            try:
                cleaned = str(val).strip().replace("$", "").replace(",", "")
                d = Decimal(cleaned)
                if d < 0:
                    raise ValueError()
                return d
            except Exception:
                raise ValueError(f"Invalid value in column {col_name}")

        headers = [
            "Sr. No.", "Name", "Designation", "Month days", "Worked days", 
            "Weekend", "CL", "Extra", "Payable Days", "Month Salary", 
            "Payable Salary", "Extra days working", "Fine/Advance", "Net Payable", 
            "Bank A/c No.", "Bank"
        ]

        total_records = 0
        success_count = 0
        failed_count = 0
        failed_rows_data = []

        try:
            for r_idx in range(2, ws.max_row + 1):
                    name_val = ws.cell(row=r_idx, column=2).value
                    if name_val is None or str(name_val).strip() == "":
                        continue

                    total_records += 1
                    row_vals = [ws.cell(row=r_idx, column=c).value for c in range(1, 17)]
                    name_str = str(name_val).strip()

                    try:
                        with transaction.atomic():
                        # Validate Employee
                            bitrix_users = BitrixClient.get_all_users()
                            employee = None
                            for u in bitrix_users:
                                mock_emp = BitrixEmployeeMock(u)
                                if mock_emp.name.lower().strip() == name_str.lower():
                                    employee = mock_emp
                                    break
                            if not employee:
                                raise ValueError("Employee not found in Bitrix24")
                            if employee.status != 'Exited':
                                raise ValueError("Cannot import active employee to dismissed payroll")

                            # Parse data
                            month_days = parse_decimal(ws.cell(row=r_idx, column=4).value, "Month days")
                            worked_days = parse_decimal(ws.cell(row=r_idx, column=5).value, "Worked days")
                            weekend = parse_decimal(ws.cell(row=r_idx, column=6).value, "Weekend")
                            cl = parse_decimal(ws.cell(row=r_idx, column=7).value, "CL")
                            extra = parse_decimal(ws.cell(row=r_idx, column=8).value, "Extra")
                            payable_days = parse_decimal(ws.cell(row=r_idx, column=9).value, "Payable Days")
                            month_salary = parse_decimal(ws.cell(row=r_idx, column=10).value, "Month Salary")
                            payable_salary = parse_decimal(ws.cell(row=r_idx, column=11).value, "Payable Salary")
                            extra_days_working = parse_decimal(ws.cell(row=r_idx, column=12).value, "Extra days working")
                            fine_advance = parse_decimal(ws.cell(row=r_idx, column=13).value, "Fine/Advance")
                            net_payable = parse_decimal(ws.cell(row=r_idx, column=14).value, "Net Payable")
                        
                            bank_account_val = ws.cell(row=r_idx, column=15).value
                            if bank_account_val is not None:
                                if isinstance(bank_account_val, float):
                                    bank_account_no = str(int(bank_account_val)).strip()
                                else:
                                    bank_account_no = str(bank_account_val).strip()
                            else:
                                bank_account_no = ""
                            
                            bank_name = str(ws.cell(row=r_idx, column=16).value or "").strip()

                            from salary.models import DismissedEmployeeBankDetail
                            # Prefill from DismissedEmployeeBankDetail if not provided in the excel row
                            if not bank_account_no or not bank_name:
                                detail = DismissedEmployeeBankDetail.objects.filter(bitrix_user_id=employee.bitrix_id).first()
                                if detail:
                                    if not bank_account_no:
                                        bank_account_no = detail.bank_account_no or ""
                                    if not bank_name:
                                        bank_name = detail.bank_name or ""

                            # If both are provided, save or update DismissedEmployeeBankDetail
                            if bank_account_no and bank_name:
                                DismissedEmployeeBankDetail.objects.update_or_create(
                                    bitrix_user_id=employee.bitrix_id,
                                    defaults={
                                        'bank_account_no': bank_account_no,
                                        'bank_name': bank_name
                                    }
                                )

                            # Save DismissedSalarySlip
                            slip = DismissedSalarySlip.objects.filter(bitrix_user_id=employee.bitrix_id, month=month, year=year).first()
                            if slip:
                                slip.month_days = month_days
                                slip.worked_days = worked_days
                                slip.weekend = weekend
                                slip.cl = cl
                                slip.extra = extra
                                slip.payable_days = payable_days
                                slip.month_salary = month_salary
                                slip.payable_salary = payable_salary
                                slip.extra_days_working = extra_days_working
                                slip.fine_advance = fine_advance
                                slip.net_payable = net_payable
                                slip.bank_account_no = bank_account_no
                                slip.bank_name = bank_name
                                slip.status = 'draft' # Re-review needed
                                slip.uploaded_batch = batch
                                slip._skip_recalculation = True
                                slip.save()
                            else:
                                slip = DismissedSalarySlip(
                                    bitrix_user_id=employee.bitrix_id,
                                    month=month,
                                    year=year,
                                    month_days=month_days,
                                    worked_days=worked_days,
                                    weekend=weekend,
                                    cl=cl,
                                    extra=extra,
                                    payable_days=payable_days,
                                    month_salary=month_salary,
                                    payable_salary=payable_salary,
                                    extra_days_working=extra_days_working,
                                    fine_advance=fine_advance,
                                    net_payable=net_payable,
                                    bank_account_no=bank_account_no,
                                    bank_name=bank_name,
                                    status='draft',
                                    uploaded_batch=batch
                                )
                                slip._skip_recalculation = True
                                slip.save()
                        
                            generate_dismissed_payslip_pdf(slip)
                            success_count += 1
                    except Exception as e:
                        failed_count += 1
                        failed_rows_data.append((row_vals, str(e)))

            if failed_count == total_records and total_records > 0:
                raise Exception("All rows failed validation")
        except Exception as batch_err:
            # Batch fully failed rollback occurred
            batch.status = 'failed'
            batch.total_records = total_records
            batch.success_count = 0
            batch.failed_count = total_records
            
            error_report_url = self.generate_error_xlsx(batch.id, headers, failed_rows_data)
            batch.error_report_path = error_report_url
            batch.save()
            return Response({
                "total": total_records,
                "success": 0,
                "failed": total_records,
                "error_report_url": error_report_url,
                "detail": str(batch_err)
            }, status=status.HTTP_400_BAD_REQUEST)

        # Batch committed
        if failed_count > 0:
            batch.status = 'partial'
            error_report_url = self.generate_error_xlsx(batch.id, headers, failed_rows_data)
            batch.error_report_path = error_report_url
            
            # Trigger 21: Salary Import Completed with Errors
            try:
                from notifications.models import Notification
                from django.contrib.auth import get_user_model
                from django.core.mail import send_mail
                
                User = get_user_model()
                admins = User.objects.filter(role__code='ADMIN')
                hrs = User.objects.filter(role__code='HR')
                
                msg = f"Salary import completed. {success_count} rows imported successfully. {failed_count} rows failed. Download error report to review."
                
                for hr in hrs:
                    Notification.objects.create(
                        recipient=hr,
                        notif_type='WARNING',
                        message=msg,
                        link="/salary/import/result/"
                    )
                    if hr.email:
                        error_full_url = f"{settings.FRONTEND_URL}{error_report_url}" if hasattr(settings, 'FRONTEND_URL') else f"http://localhost:8000{error_report_url}"
                        send_mail(
                            subject="Salary Import Completed with Errors",
                            message=f"{msg}\n\nDownload the error report here: {error_full_url}",
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            recipient_list=[hr.email],
                            fail_silently=True
                        )
                for admin in admins:
                    Notification.objects.create(
                        recipient=admin,
                        notif_type='WARNING',
                        message=msg,
                        link="/salary/import/result/"
                    )
            except Exception:
                pass
        else:
            batch.status = 'success'
            error_report_url = None

        batch.total_records = total_records
        batch.success_count = success_count
        batch.failed_count = failed_count
        batch.save()

        return Response({
            "total": total_records,
            "success": success_count,
            "failed": failed_count,
            "error_report_url": error_report_url
        }, status=status.HTTP_200_OK)

    def generate_error_xlsx(self, batch_id, headers, failed_rows_data):
        error_report_dir = os.path.join(settings.MEDIA_ROOT, "error_reports")
        os.makedirs(error_report_dir, exist_ok=True)
        report_filename = f"error_report_{batch_id}_{int(datetime.datetime.now().timestamp())}.xlsx"
        report_path = os.path.join(error_report_dir, report_filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Errors"
        ws.append(headers + ["Error Reason"])

        for row_vals, reason in failed_rows_data:
            ws.append(row_vals + [reason])

        wb.save(report_path)
        return f"{settings.MEDIA_URL}error_reports/{report_filename}"


class DismissedSalaryPublishView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can publish slips.")

        month_str = request.data.get('month')
        year_str = request.data.get('year')

        if not month_str or not year_str:
            return Response({'error': 'month and year are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month = int(month_str)
            year = int(year_str)
        except ValueError:
            return Response({'error': 'Invalid month or year.'}, status=status.HTTP_400_BAD_REQUEST)

        slips = DismissedSalarySlip.objects.filter(month=month, year=year, status='draft')
        count = slips.count()
        for slip in slips:
            slip.status = 'published'
            slip.save()
            generate_dismissed_payslip_pdf(slip)

        # Trigger 20 notification
        import calendar
        try:
            month_name = calendar.month_name[month]
            month_year = f"{month_name} {year}"
            month_year_slug = f"{month_name.lower()}-{year}"
            
            from notifications.models import Notification
            from django.contrib.auth import get_user_model
            User = get_user_model()
            admins_and_hrs = User.objects.filter(role__code__in=['ADMIN', 'HR'])
            msg = f"Salary slips generated for {month_year}. {count} slips ready for download."
            for user in admins_and_hrs:
                Notification.objects.create(
                    recipient=user,
                    notif_type='SUCCESS',
                    message=msg,
                    link=f"/salary/slips/{month_year_slug}/"
                )
        except Exception:
            pass

        return Response({'message': f'Successfully published {count} salary slips.'}, status=status.HTTP_200_OK)


class DismissedSalaryEditView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can edit salary slips.")

        slip = get_object_or_404(DismissedSalarySlip, pk=pk)

        # Exclude read-only calculations from direct parsing
        mutable_fields = [
            'location', 'month_days', 'worked_days', 'weekend', 'cl', 'extra',
            'month_salary', 'extra_days_working', 'fine_advance',
            'bank_account_no', 'bank_name',
            'leaves_available', 'working_days', 'extra_days',
            'gross_salary', 'pf_contribution', 'esi', 'labour_welfare_fund',
            'professional_tax', 'other_deductions', 'net_credited_amount',
            'payment_status', 'payment_date', 'transaction_ref'
        ]

        for field in mutable_fields:
            if field in request.data:
                val = request.data[field]
                if field in ['payment_status', 'transaction_ref', 'location', 'bank_account_no', 'bank_name']:
                    setattr(slip, field, val)
                elif field == 'payment_date':
                    setattr(slip, field, val if val else None)
                else:
                    setattr(slip, field, Decimal(str(val)) if val else Decimal('0.00'))

        slip.status = 'draft' # Re-review needed after edit
        slip.save()

        # Update or create persistent employee bank details if both are present in the slip
        if slip.bank_account_no and slip.bank_name:
            from salary.models import DismissedEmployeeBankDetail
            DismissedEmployeeBankDetail.objects.update_or_create(
                bitrix_user_id=slip.bitrix_user_id,
                defaults={
                    'bank_account_no': slip.bank_account_no,
                    'bank_name': slip.bank_name
                }
            )

        generate_dismissed_payslip_pdf(slip)

        serializer = DismissedSalarySlipSerializer(slip)
        return Response(serializer.data, status=status.HTTP_200_OK)


def get_employee_slips_for_range_dismissed(employee_id, from_year, from_month, to_year, to_month, status_filter=None, payment_status=None):
    # 1. Fetch real slips
    real_slips = DismissedSalarySlip.objects.filter(bitrix_user_id=str(employee_id))
    if status_filter:
        real_slips = real_slips.filter(status=status_filter)
    if payment_status:
        real_slips = real_slips.filter(payment_status=payment_status)

    # Keep track of which periods have real slips
    real_periods = set()
    slips_list = []
    for s in real_slips:
        real_periods.add((s.year, s.month))
        slips_list.append(s)

    imported_periods = set(
        DismissedSalarySlip.objects.filter(bitrix_user_id=str(employee_id))
        .values_list('year', 'month')
    )

    # Determine bounds for date range
    first_slip = DismissedSalarySlip.objects.filter(bitrix_user_id=str(employee_id)).order_by('year', 'month').first()
    
    # Cap carry-forward up to current local date
    from django.utils import timezone
    today = timezone.localdate()
    cap_Y, cap_M = today.year, today.month

    # Sort out bounds
    fY, fM = from_year, from_month
    if fY is None and first_slip:
        fY, fM = first_slip.year, first_slip.month
    elif fY is None:
        fY, fM = 2000, 1

    tY, tM = to_year, to_month
    if tY is None:
        tY, tM = cap_Y, cap_M

    # Cap range end for carry-forward mock generation to current local date
    gen_to_Y, gen_to_M = tY, tM
    if gen_to_Y > cap_Y or (gen_to_Y == cap_Y and gen_to_M > cap_M):
        gen_to_Y, gen_to_M = cap_Y, cap_M

    # Iterate month-by-month and generate carry-forward slips
    curr_year, curr_month = fY, fM
    # if first_slip:
    #     start_year, start_month = first_slip.year, first_slip.month
    #     while (curr_year < gen_to_Y) or (curr_year == gen_to_Y and curr_month <= gen_to_M):
    #         if (curr_year > start_year) or (curr_year == start_year and curr_month >= start_month):
    #             if (curr_year, curr_month) not in imported_periods:
    #                 prior_slip = get_latest_prior_dismissed_slip(employee_id, curr_month, curr_year)
    #                 if prior_slip:
    #                     if not status_filter or prior_slip.status == status_filter:
    #                         if not payment_status or prior_slip.payment_status == payment_status:
    #                             mock_slip = calculate_carry_forward_dismissed_slip(employee_id, curr_month, curr_year, prior_slip)
    #                             slips_list.append(mock_slip)
    #         # Increment month
    #         curr_month += 1
    #         if curr_month > 12:
    #             curr_month = 1
    #             curr_year += 1

    # Filter and sort
    final_slips = []
    for s in slips_list:
        if fY is not None:
            if s.year < fY or (s.year == fY and s.month < fM):
                continue
        if tY is not None:
            if s.year > tY or (s.year == tY and s.month > tM):
                continue
        final_slips.append(s)
        
    final_slips.sort(key=lambda s: (s.year, s.month), reverse=True)
    return final_slips


class DismissedSalaryHistoryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        role = check_role(request.user)
        
        # Scoped logic
        employee_id = None
        if role == 'employee':
            user_data = next((u for u in BitrixClient.get_all_users() if u.get('email') == request.user.email), None)
            if not user_data:
                return Response({'error': 'Employee profile not found in Bitrix24.'}, status=status.HTTP_404_NOT_FOUND)
            employee = BitrixEmployeeMock(user_data)
            employee_id = employee.bitrix_id
        else:
            employee_id = request.query_params.get('employee_id') or kwargs.get('employee_id') or request.query_params.get('bitrix_user_id')

        # Payment status filtering
        payment_status = request.query_params.get('payment_status')

        # Date range filtering
        from_param = request.query_params.get('from') # format: YYYY-MM
        to_param = request.query_params.get('to')     # format: YYYY-MM

        fY, fM = None, None
        if from_param:
            try:
                fY, fM = map(int, from_param.split('-'))
            except ValueError:
                pass

        tY, tM = None, None
        if to_param:
            try:
                tY, tM = map(int, to_param.split('-'))
            except ValueError:
                pass

        if employee_id:
            status_filter = 'published' if role == 'employee' else None
            slips = get_employee_slips_for_range_dismissed(
                employee_id=employee_id,
                from_year=fY,
                from_month=fM,
                to_year=tY,
                to_month=tM,
                status_filter=status_filter,
                payment_status=payment_status
            )
        else:
            # Admin / HR paginating all employees
            slips = DismissedSalarySlip.objects.all().order_by('-year', '-month')
            if payment_status:
                slips = slips.filter(payment_status=payment_status)
            if fY is not None:
                slips = slips.filter(models.Q(year__gt=fY) | models.Q(year=fY, month__gte=fM))
            if tY is not None:
                slips = slips.filter(models.Q(year__lt=tY) | models.Q(year=tY, month__lte=tM))

        # Pagination
        paginator = StandardResultsSetPagination()
        page = paginator.paginate_queryset(slips, request, view=self)
        
        bitrix_users = {str(u['id']): BitrixEmployeeMock(u) for u in BitrixClient.get_all_users()}
        data = []
        target_list = page if page is not None else slips
        for s in target_list:
            emp = bitrix_users.get(str(s.bitrix_user_id))
            emp_id = emp.id if emp else 0
            emp_name = emp.name if emp else f"User {s.bitrix_user_id}"
            data.append({
                'id': s.id,
                'employee_id': emp_id,
                'employee_name': emp_name,
                'month': s.month,
                'year': s.year,
                'location': s.location,
                'leaves_available': str(s.leaves_available),
                'working_days': str(s.working_days),
                'extra_days': str(s.extra_days),
                'gross_salary': str(s.gross_salary),
                'pf_contribution': str(s.pf_contribution),
                'esi': str(s.esi),
                'labour_welfare_fund': str(s.labour_welfare_fund),
                'professional_tax': str(s.professional_tax),
                'other_deductions': str(s.other_deductions),
                'total_deductions': str(s.total_deductions),
                'net_salary': str(s.net_salary),
                'net_credited_amount': str(s.net_credited_amount),
                'month_days': str(s.month_days),
                'worked_days': str(s.worked_days),
                'weekend': str(s.weekend),
                'cl': str(s.cl),
                'extra': str(s.extra),
                'payable_days': str(s.payable_days),
                'month_salary': str(s.month_salary),
                'payable_salary': str(s.payable_salary),
                'extra_days_working': str(s.extra_days_working),
                'fine_advance': str(s.fine_advance),
                'net_payable': str(s.net_payable),
                'bank_account_no': s.bank_account_no or "",
                'bank_name': s.bank_name or "",
                'payment_status': s.payment_status,
                'payment_date': str(s.payment_date) if s.payment_date else '',
                'transaction_ref': s.transaction_ref or '',
                'status': s.status
            })

        if page is not None:
            return paginator.get_paginated_response(data)
        return Response(data)


class DismissedSalarySlipDownloadView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        role = check_role(request.user)

        employee_id = request.query_params.get('employee_id')
        download_type = request.query_params.get('type') # single | last3 | last4 | range | bulk_month | selected

        if role == 'employee':
            user_data = next((u for u in BitrixClient.get_all_users() if u.get('email') == request.user.email), None)
            if not user_data:
                return Response({'error': 'Employee profile not found.'}, status=status.HTTP_404_NOT_FOUND)
            employee = BitrixEmployeeMock(user_data)
            employee_id = employee.bitrix_id
            # Employees can only download published slips
            slips_qs = DismissedSalarySlip.objects.filter(bitrix_user_id=employee_id, status='published')
        else:
            # Admin/HR
            if download_type in ['bulk_month', 'selected'] and not employee_id:
                slips_qs = DismissedSalarySlip.objects.all()
            else:
                if not employee_id:
                    return Response({'error': 'employee_id parameter is required for Admin/HR.'}, status=status.HTTP_400_BAD_REQUEST)
                slips_qs = DismissedSalarySlip.objects.filter(bitrix_user_id=str(employee_id))

        # Filters
        slips = []
        if download_type == 'single':
            month_str = request.query_params.get('month')
            year_str = request.query_params.get('year')
            if not month_str or not year_str:
                return Response({'error': 'month and year are required.'}, status=status.HTTP_400_BAD_REQUEST)
            slips = slips_qs.filter(month=int(month_str), year=int(year_str))
            
            slips = list(slips)
            if not slips and employee_id:
                # Carry forward logic!
                month = int(month_str)
                year = int(year_str)
                prior_slip = get_latest_prior_dismissed_slip(employee_id, month, year)
                if prior_slip:
                    if role != 'employee' or prior_slip.status == 'published':
                        carry_slip = calculate_carry_forward_dismissed_slip(employee_id, month, year, prior_slip)
                        pdf_bytes, filename = generate_dismissed_payslip_pdf_bytes(carry_slip)
                        response = HttpResponse(pdf_bytes, content_type='application/pdf')
                        response['Content-Disposition'] = f'attachment; filename="{filename}"'
                        return response
        elif download_type == 'last3':
            slips = slips_qs.order_by('-year', '-month')[:3]
            slips = list(slips)
        elif download_type == 'last4':
            slips = slips_qs.order_by('-year', '-month')[:4]
            slips = list(slips)
        elif download_type == 'range':
            from_month = int(request.query_params.get('from_month', 1))
            from_year = int(request.query_params.get('from_year', 2000))
            to_month = int(request.query_params.get('to_month', 12))
            to_year = int(request.query_params.get('to_year', 2100))
            
            if not employee_id:
                return Response({'error': 'employee_id parameter is required.'}, status=status.HTTP_400_BAD_REQUEST)
            
            status_filter = 'published' if role == 'employee' else None
            slips = get_employee_slips_for_range_dismissed(
                employee_id=employee_id,
                from_year=from_year,
                from_month=from_month,
                to_year=to_year,
                to_month=to_month,
                status_filter=status_filter
            )
        elif download_type == 'bulk_month':
            if role not in ['admin', 'hr']:
                raise PermissionDenied("Only Admin/HR can bulk download slips.")
            month_str = request.query_params.get('month')
            year_str = request.query_params.get('year')
            if not month_str or not year_str:
                return Response({'error': 'month and year are required.'}, status=status.HTTP_400_BAD_REQUEST)
            slips = slips_qs.filter(month=int(month_str), year=int(year_str))
            slips = list(slips)
        elif download_type == 'selected':
            slip_ids_str = request.query_params.get('slip_ids')
            if not slip_ids_str:
                return Response({'error': 'slip_ids parameter is required for type=selected.'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                ids = [int(x) for x in slip_ids_str.split(',')]
                slips = slips_qs.filter(id__in=ids)
                slips = list(slips)
            except ValueError:
                return Response({'error': 'Invalid slip_ids parameter.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Invalid download type.'}, status=status.HTTP_400_BAD_REQUEST)

        if not slips:
            return Response({'error': 'No salary slips found matching parameters.'}, status=status.HTTP_404_NOT_FOUND)

        if len(slips) == 1:
            slip = slips[0]
            # Ensure PDF exists (only for real database records, mock ones don't have id/pdf_file field saved)
            if getattr(slip, 'id', None) is not None:
                generate_dismissed_payslip_pdf(slip)
                response = HttpResponse(slip.pdf_file.read(), content_type='application/pdf')
                emp_id_str = slip.employee.emp_id if slip.employee else slip.bitrix_user_id
                response['Content-Disposition'] = f'attachment; filename="payslip_{emp_id_str}_{slip.month}_{slip.year}.pdf"'
                return response
            else:
                # Mock slip single download
                pdf_bytes, filename = generate_dismissed_payslip_pdf_bytes(slip)
                response = HttpResponse(pdf_bytes, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
        else:
            # Serve ZIP
            zip_type = 'bulk' if download_type == 'bulk_month' else 'employee'
            zip_data = generate_dismissed_payslips_zip(slips, zip_type)
            
            response = HttpResponse(zip_data, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="payslips_{int(datetime.datetime.now().timestamp())}.zip"'
            return response


class DismissedSalaryImportBatchesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can view import batches.")

        batches = DismissedSalaryImportBatch.objects.all().order_by('-uploaded_at')
        paginator = StandardResultsSetPagination()
        page = paginator.paginate_queryset(batches, request, view=self)
        serializer = DismissedSalaryImportBatchSerializer(page if page is not None else batches, many=True)
        
        if page is not None:
            return paginator.get_paginated_response(serializer.data)
        return Response(serializer.data)


class DismissedSalaryEmployeeSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, employee_id):
        role = check_role(request.user)
        if role == 'employee':
            # Employees can only view their own summary
            user_data = next((u for u in BitrixClient.get_all_users() if u.get('email') == request.user.email), None)
            if not user_data or str(user_data.get('id')) != str(employee_id):
                raise PermissionDenied("You can only view your own salary summary.")

        slips = DismissedSalarySlip.objects.filter(bitrix_user_id=str(employee_id))

        # Payment status filtering
        payment_status = request.query_params.get('payment_status')
        if payment_status:
            slips = slips.filter(payment_status=payment_status)

        # Date range filtering
        from_param = request.query_params.get('from') # format: YYYY-MM
        to_param = request.query_params.get('to')     # format: YYYY-MM

        if from_param:
            try:
                fY, fM = map(int, from_param.split('-'))
                slips = slips.filter(
                    models.Q(year__gt=fY) | models.Q(year=fY, month__gte=fM)
                )
            except ValueError:
                pass

        if to_param:
            try:
                tY, tM = map(int, to_param.split('-'))
                slips = slips.filter(
                    models.Q(year__lt=tY) | models.Q(year=tY, month__lte=tM)
                )
            except ValueError:
                pass
        
        total_credited = sum((slip.net_payable or Decimal('0.00') for slip in slips), Decimal('0.00'))
        total_deductions = sum((slip.fine_advance or Decimal('0.00') for slip in slips), Decimal('0.00'))
        total_payslips = slips.count()
        last_paid_slip = slips.filter(payment_status='paid', payment_date__isnull=False).order_by('-payment_date').first()
        last_payment_date = last_paid_slip.payment_date.strftime('%Y-%m-%d') if last_paid_slip else '-'

        return Response({
            'total_salary_credited': str(total_credited),
            'total_deductions': str(total_deductions),
            'total_payslips': total_payslips,
            'last_payment_date': last_payment_date
        })


class DismissedSalaryEmployeeHistoryExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, employee_id):
        role = check_role(request.user)
        if role == 'employee':
            user_data = next((u for u in BitrixClient.get_all_users() if u.get('email') == request.user.email), None)
            if not user_data or str(user_data.get('id')) != str(employee_id):
                raise PermissionDenied("You can only export your own salary history.")
        else:
            user_data = BitrixClient.get_user_detail(employee_id)

        if not user_data:
            return Response({'error': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        employee = BitrixEmployeeMock(user_data)
        slips = DismissedSalarySlip.objects.filter(bitrix_user_id=employee.bitrix_id).order_by('-year', '-month')

        from_param = request.query_params.get('from')
        to_param = request.query_params.get('to')
        payment_status = request.query_params.get('payment_status')

        if from_param:
            try:
                fY, fM = map(int, from_param.split('-'))
                slips = slips.filter(models.Q(year__gt=fY) | models.Q(year=fY, month__gte=fM))
            except ValueError:
                pass
        if to_param:
            try:
                tY, tM = map(int, to_param.split('-'))
                slips = slips.filter(models.Q(year__lt=tY) | models.Q(year=tY, month__lte=tM))
            except ValueError:
                pass
        if payment_status:
            slips = slips.filter(payment_status=payment_status)

        headers = [
            "Month/Year", "Location", "Month days", "Worked days", "Weekend", 
            "CL", "Extra", "Payable Days", "Month Salary", "Payable Salary", 
            "Extra days working", "Fine/Advance", "Net Payable", "Bank A/c No.", "Bank",
            "Payment Status", "Payment Date", "Transaction Ref"
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Salary History"
        ws.append(headers)

        month_names = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

        for s in slips:
            ws.append([
                f"{month_names[s.month]} {s.year}",
                s.location,
                float(s.month_days),
                float(s.worked_days),
                float(s.weekend),
                float(s.cl),
                float(s.extra),
                float(s.payable_days),
                float(s.month_salary),
                float(s.payable_salary),
                float(s.extra_days_working),
                float(s.fine_advance),
                float(s.net_payable),
                s.bank_account_no or "",
                s.bank_name or "",
                s.payment_status.capitalize(),
                s.payment_date.strftime('%Y-%m-%d') if s.payment_date else '',
                s.transaction_ref or ''
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="salary_history_{employee.emp_id}.xlsx"'
        wb.save(response)
        return response


class DismissedSalaryIndividualGenerateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can manually generate payslips.")

        employee_id = request.data.get('employee_id')
        month_str = request.data.get('month')
        year_str = request.data.get('year')

        if not employee_id or not month_str or not year_str:
            return Response({'error': 'employee_id, month, and year are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month = int(month_str)
            year = int(year_str)
            if not (1 <= month <= 12):
                raise ValueError()
        except ValueError:
            return Response({'error': 'Invalid month or year.'}, status=status.HTTP_400_BAD_REQUEST)

        # Check if an imported slip already exists
        existing_imported_slip = DismissedSalarySlip.objects.filter(bitrix_user_id=str(employee_id), month=month, year=year, uploaded_batch__isnull=False).first()
        if existing_imported_slip:
            generate_dismissed_payslip_pdf(existing_imported_slip)
            return Response({'message': 'Payslip generated successfully.', 'id': existing_imported_slip.id}, status=status.HTTP_200_OK)

        # Get existing non-imported slip if any
        existing_slip = DismissedSalarySlip.objects.filter(bitrix_user_id=str(employee_id), month=month, year=year).first()
        is_update = existing_slip is not None

        # 1. Try to copy from latest prior slip
        prior_slip = get_latest_prior_dismissed_slip(employee_id, month, year)
        if prior_slip:
            slip_data = calculate_carry_forward_dismissed_slip(employee_id, month, year, prior_slip)
            if existing_slip:
                # Update existing non-imported slip with carry forward values
                existing_slip.location = slip_data.location
                existing_slip.month_days = slip_data.month_days
                existing_slip.worked_days = slip_data.worked_days
                existing_slip.weekend = slip_data.weekend
                existing_slip.cl = slip_data.cl
                existing_slip.extra = slip_data.extra
                existing_slip.payable_days = slip_data.payable_days
                existing_slip.month_salary = slip_data.month_salary
                existing_slip.payable_salary = slip_data.payable_salary
                existing_slip.extra_days_working = slip_data.extra_days_working
                existing_slip.fine_advance = slip_data.fine_advance
                existing_slip.net_payable = slip_data.net_payable
                existing_slip.bank_account_no = slip_data.bank_account_no
                existing_slip.bank_name = slip_data.bank_name
                existing_slip.status = 'published'
                existing_slip._skip_recalculation = True
                existing_slip.save()
                slip = existing_slip
            else:
                slip_data.status = 'published'
                slip_data._skip_recalculation = True
                slip_data.save()
                slip = slip_data
        else:
            # 2. Try to generate from active DismissedSalaryStructure
            from salary.models import DismissedSalaryStructure, DismissedEmployeeBankDetail
            structure = DismissedSalaryStructure.objects.filter(bitrix_user_id=str(employee_id)).order_by('-effective_from').first()
            if not structure:
                return Response({'error': 'No prior salary slip or salary structure found for this employee to carry forward from.'}, status=status.HTTP_400_BAD_REQUEST)
            
            bank_detail = DismissedEmployeeBankDetail.objects.filter(bitrix_user_id=str(employee_id)).first()
            bank_acc = bank_detail.bank_account_no if bank_detail else ""
            bank_name = bank_detail.bank_name if bank_detail else ""
            
            # Use month days
            import calendar
            month_days = Decimal(calendar.monthrange(year, month)[1])
            
            if existing_slip:
                existing_slip.location = 'Mohali'
                existing_slip.month_days = month_days
                existing_slip.worked_days = month_days
                existing_slip.weekend = Decimal('0.00')
                existing_slip.cl = Decimal('0.00')
                existing_slip.extra = Decimal('0.00')
                existing_slip.payable_days = month_days
                existing_slip.month_salary = structure.gross_salary
                existing_slip.payable_salary = structure.gross_salary
                existing_slip.extra_days_working = Decimal('0.00')
                existing_slip.fine_advance = Decimal('0.00')
                existing_slip.net_payable = structure.net_salary
                existing_slip.bank_account_no = bank_acc
                existing_slip.bank_name = bank_name
                existing_slip.status = 'published'
                existing_slip._skip_recalculation = True
                existing_slip.save()
                slip = existing_slip
            else:
                slip = DismissedSalarySlip(
                    bitrix_user_id=str(employee_id),
                    month=month,
                    year=year,
                    location='Mohali',
                    month_days=month_days,
                    worked_days=month_days,
                    weekend=Decimal('0.00'),
                    cl=Decimal('0.00'),
                    extra=Decimal('0.00'),
                    payable_days=month_days,
                    month_salary=structure.gross_salary,
                    payable_salary=structure.gross_salary,
                    extra_days_working=Decimal('0.00'),
                    fine_advance=Decimal('0.00'),
                    net_payable=structure.net_salary,
                    bank_account_no=bank_acc,
                    bank_name=bank_name,
                    payment_status='pending',
                    status='published'
                )
                slip._skip_recalculation = True
                slip.save()

        generate_dismissed_payslip_pdf(slip)
        return Response({'message': 'Payslip generated successfully.', 'id': slip.id}, status=status.HTTP_200_OK if is_update else status.HTTP_201_CREATED)

class DismissedSalaryGridView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can access salary grid.")

        month_param = request.query_params.get('month')
        year_param = request.query_params.get('year')

        if not month_param or not year_param:
            return Response({'error': 'month and year parameters are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month = int(month_param)
            year = int(year_param)
        except ValueError:
            return Response({'error': 'Invalid month or year.'}, status=status.HTTP_400_BAD_REQUEST)

        slips = DismissedSalarySlip.objects.filter(month=month, year=year)
        slips_map = {str(slip.bitrix_user_id): slip for slip in slips}
        
        user_map = {}
        for u in BitrixClient.get_all_users():
            user_map[str(u['id'])] = BitrixEmployeeMock(u)
            
        from salary.models import DismissedEmployeeBankDetail
        import calendar
        from exit_formality.models import ExitRequest

        grid_data = []

        month_days_val = calendar.monthrange(year, month)[1]
        
        employees = []
        for emp in user_map.values():
            if emp.status == 'Exited':
                employees.append(emp)
        
        for emp in employees:
            slip = slips_map.get(str(emp.bitrix_id))
            
            if slip:
                emp_name = emp.name if emp else f"User {slip.bitrix_user_id}"
                designation = emp.designation if emp else ""
                
                detail = DismissedEmployeeBankDetail.objects.filter(bitrix_user_id=slip.bitrix_user_id).first()
                bank_acc = detail.bank_account_no if (detail and detail.bank_account_no) else (slip.bank_account_no or "")
                bank_nm = detail.bank_name if (detail and detail.bank_name) else (slip.bank_name or "")

                grid_data.append({
                    "id": slip.id,
                    "bitrix_id": emp.bitrix_id,
                    "name": emp_name,
                    "designation": designation,
                    "month_days": float(slip.month_days),
                    "worked_days": float(slip.worked_days),
                    "weekend": float(slip.weekend),
                    "cl": float(slip.cl),
                    "extra": float(slip.extra),
                    "payable_days": float(slip.payable_days),
                    "month_salary": float(slip.month_salary),
                    "payable_salary": float(slip.payable_salary),
                    "extra_days_working": float(slip.extra_days_working),
                    "fine_advance": float(slip.fine_advance),
                    "net_payable": float(slip.net_payable),
                    "bank_account_no": bank_acc,
                    "bank_name": bank_nm
                })
            else:
                prior_slip = get_latest_prior_dismissed_slip(emp.bitrix_id, month, year)
                
                struct = None
                if not prior_slip:
                    import datetime
                    struct = DismissedSalaryStructure.objects.filter(
                        bitrix_user_id=emp.bitrix_id,
                        effective_from__lte=datetime.date(year, month, 28)
                    ).order_by('-effective_from').first()
                    if not struct:
                        struct = DismissedSalaryStructure.objects.filter(
                            bitrix_user_id=emp.bitrix_id
                        ).order_by('-effective_from').first()

                detail = DismissedEmployeeBankDetail.objects.filter(bitrix_user_id=emp.bitrix_id).first()
                bank_acc = (detail.bank_account_no if detail and detail.bank_account_no else None) or \
                           (prior_slip.bank_account_no if prior_slip else None) or \
                           getattr(emp, 'bank_account', '') or ""
                bank_nm = (detail.bank_name if detail and detail.bank_name else None) or \
                          (prior_slip.bank_name if prior_slip else None) or ""

                if prior_slip:
                    temp_slip = calculate_carry_forward_dismissed_slip(emp.bitrix_id, month, year, prior_slip)
                    grid_data.append({
                        "id": None,
                        "bitrix_id": emp.bitrix_id,
                        "name": emp.name,
                        "designation": emp.designation,
                        "month_days": float(temp_slip.month_days),
                        "worked_days": float(temp_slip.worked_days),
                        "weekend": float(temp_slip.weekend),
                        "cl": float(temp_slip.cl),
                        "extra": float(temp_slip.extra),
                        "payable_days": float(temp_slip.payable_days),
                        "month_salary": float(temp_slip.month_salary),
                        "payable_salary": float(temp_slip.payable_salary),
                        "extra_days_working": float(temp_slip.extra_days_working),
                        "fine_advance": float(temp_slip.fine_advance),
                        "net_payable": float(temp_slip.net_payable),
                        "bank_account_no": bank_acc,
                        "bank_name": bank_nm
                    })
                else:
                    grid_data.append({
                        "id": None,
                        "bitrix_id": emp.bitrix_id,
                        "name": emp.name,
                        "designation": emp.designation,
                        "month_days": float(month_days_val),
                        "worked_days": 0.0,
                        "weekend": 0.0,
                        "cl": 0.0,
                        "extra": 0.0,
                        "payable_days": 0.0,
                        "month_salary": float(struct.monthly_gross) if struct else 0.0,
                        "payable_salary": 0.0,
                        "extra_days_working": 0.0,
                        "fine_advance": 0.0,
                        "net_payable": 0.0,
                        "bank_account_no": bank_acc,
                        "bank_name": bank_nm
                    })

        return Response(grid_data)

    def post(self, request):
        if check_role(request.user) != 'admin':
            raise PermissionDenied("Only Admin can save salary grid.")

        month_str = request.data.get('month')
        year_str = request.data.get('year')
        rows = request.data.get('rows', [])

        if not month_str or not year_str:
            return Response({'error': 'month and year are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month = int(month_str)
            year = int(year_str)
        except ValueError:
            return Response({'error': 'Invalid month or year.'}, status=status.HTTP_400_BAD_REQUEST)

        def parse_decimal(val):
            if val is None or str(val).strip() == "":
                return Decimal("0.00")
            try:
                cleaned = str(val).strip().replace("$", "").replace(",", "")
                d = Decimal(cleaned)
                if d < 0:
                    raise ValueError()
                return d
            except Exception:
                return Decimal("0.00")

        total_records = 0
        saved_slips = []
        with transaction.atomic():
            for row in rows:
                bitrix_id = str(row.get('bitrix_id', ''))
                if not bitrix_id:
                    continue

                total_records += 1

                month_days = parse_decimal(row.get('month_days'))
                worked_days = parse_decimal(row.get('worked_days'))
                weekend = parse_decimal(row.get('weekend'))
                cl = parse_decimal(row.get('cl'))
                extra = parse_decimal(row.get('extra'))
                payable_days = parse_decimal(row.get('payable_days'))
                month_salary = parse_decimal(row.get('month_salary'))
                payable_salary = parse_decimal(row.get('payable_salary'))
                extra_days_working = parse_decimal(row.get('extra_days_working'))
                fine_advance = parse_decimal(row.get('fine_advance'))
                net_payable = parse_decimal(row.get('net_payable'))
                bank_acc = row.get('bank_account_no', '')
                bank_nm = row.get('bank_name', '')

                slip = DismissedSalarySlip.objects.filter(bitrix_user_id=bitrix_id, month=month, year=year).first()
                if slip:
                    slip.month_days = month_days
                    slip.worked_days = worked_days
                    slip.weekend = weekend
                    slip.cl = cl
                    slip.extra = extra
                    slip.payable_days = payable_days
                    slip.month_salary = month_salary
                    slip.payable_salary = payable_salary
                    slip.extra_days_working = extra_days_working
                    slip.fine_advance = fine_advance
                    slip.net_payable = net_payable
                    if bank_acc: slip.bank_account_no = bank_acc
                    if bank_nm: slip.bank_name = bank_nm
                    slip.status = 'draft'
                    slip._skip_recalculation = True
                    slip.save()
                    saved_slips.append(slip.id)
                else:
                    slip = DismissedSalarySlip(
                        bitrix_user_id=bitrix_id,
                        month=month,
                        year=year,
                        month_days=month_days,
                        worked_days=worked_days,
                        weekend=weekend,
                        cl=cl,
                        extra=extra,
                        payable_days=payable_days,
                        month_salary=month_salary,
                        payable_salary=payable_salary,
                        extra_days_working=extra_days_working,
                        fine_advance=fine_advance,
                        net_payable=net_payable,
                        bank_account_no=bank_acc,
                        bank_name=bank_nm,
                        payment_status='pending',
                        status='draft'
                    )
                    slip._skip_recalculation = True
                    slip.save()
                    saved_slips.append(slip.id)

        # Generate PDFs in the background to prevent blocking the UI
        import threading
        def generate_pdfs_in_background(slip_ids):
            from salary.models import DismissedSalarySlip
            for sid in slip_ids:
                s = DismissedSalarySlip.objects.filter(id=sid).first()
                if s:
                    try:
                        generate_dismissed_payslip_pdf(s)
                    except Exception:
                        pass
                        
        threading.Thread(target=generate_pdfs_in_background, args=(saved_slips,)).start()

        return Response({'message': f'Successfully saved {total_records} records for {month}/{year}.'}, status=status.HTTP_200_OK)
